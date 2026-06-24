###########################################################
# Aquí van las bibliotecas empleadas para hacer funcionar
# mi TFG.
###########################################################

import base64
from collections import defaultdict
import io
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os # Esta librería me permite acceder a las diferentes rutas de los archivos y directorios, lo que me permite cargar el POD y exportarlo
import pandas as pandas
import pdfplumber
import re # Las expresiones regulares que importa esta librería sirven para parsear correctamente el POD
import streamlit as st
import streamlit.components.v1 as componentes

###########################################################
# Aquí está el CSS que he hecho para darle estilos a la
# página de mi TFG.
###########################################################

st.markdown("""
    <style>
    /* <<<<<<<<<<<<<<<<<<<< IMPORTACIÓN DE LA FUENTE DE TEXTO >>>>>>>>>>>>>>>>>>>> */
            
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500;600;700&family=IBM+Plex+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&display=swap');

    /* <<<<<<<<<<<<<<<<<<<< VALORES BASE Y GLOBALES >>>>>>>>>>>>>>>>>>>> */
            
    html, body, [class*="css"] {
        font-family: 'IBM Plex Sans', sans-serif;
        color: #1a1a1a;
    }

    .stApp {
        background: #eeeae4;
    }

    /* <<<<<<<<<<<<<<<<<<<< BARRA LATERAL >>>>>>>>>>>>>>>>>>>> */
            
    [data-testid="stSidebar"] {
        background: #1c1c1c !important;
        border-right: none !important;
    }

    /* Textos de la barra lateral */
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] label {
        color: #e8e4de !important;
    }

    /* Botones de la barra lateral */
            
    [data-testid="stSidebar"] .stButton>button {
        background: #2a2a2a !important;
        color: #f0ede8 !important;
        border: 1px solid #3a3a3a !important;
        border-radius: 4px !important;
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: .78rem !important;
        text-transform: uppercase !important;
        letter-spacing: .05em !important;
        margin-bottom: .2rem !important;
        width: 100% !important;
    }

    [data-testid="stSidebar"] .stButton>button:hover {
        background: #3a3a3a !important;
        color: #ffffff !important;
    }

    [data-testid="stSidebar"] .stButton>button[kind="primary"] {
        background: #e84545 !important;
        border-color: #e84545 !important;
        color: #fff !important;
    }

    /* Selectores y checkboxes de la barra lateral */
            
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stCheckbox label {
        color: #e8e4de !important;
    }

    /* Procesamiento del texto en la barra lateral por medio de Markdown */
            
    [data-testid="stSidebar"] .stMarkdown, 
    [data-testid="stSidebar"] .stMarkdown div {
        color: #e8e4de !important;
    }

    /* <<<<<<<<<<<<<<<<<<<< BOTONES GENERALES >>>>>>>>>>>>>>>>>>>> */
            
    .stButton>button {
        background: #1c1c1c;
        color: #f0ede8;
        border: none;
        border-radius: 4px;
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        font-size: .78rem;
        text-transform: uppercase;
        letter-spacing: .05em;
        padding: .45rem 1.1rem;
        transition: background .15s;
    }

    .stButton>button:hover {
        background: #333;
    }

    .stButton>button[kind="secondary"] {
        background: transparent;
        color: #1c1c1c;
        border: 1.5px solid #1c1c1c;
    }

    .stButton>button[kind="secondary"]:hover {
        background: #dedad4;
    }

    /* <<<<<<<<<<<<<<<<<<<< BOTONES DE DESCARGA >>>>>>>>>>>>>>>>>>>> */
            
    .stDownloadButton>button {
        background: transparent;
        color: #1c1c1c;
        border: 2px solid #1c1c1c;
        border-radius: 4px;
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        font-size: .78rem;
        text-transform: uppercase;
        letter-spacing: .05em;
    }

    /* <<<<<<<<<<<<<<<<<<<< ENCABEZADOS DE SECCIÓN >>>>>>>>>>>>>>>>>>>> */
            
    .sec-hdr {
        font-family: 'IBM Plex Mono', monospace;
        font-size: .66rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: .16em;
        color: #777;
        border-bottom: 1px solid #ccc;
        padding-bottom: .4rem;
        margin: 1.3rem 0 .9rem 0;
    }

    /* <<<<<<<<<<<<<<<<<<<< CAJAS DE INFORMACIÓN >>>>>>>>>>>>>>>>>>>> */
            
    .box-info {
        background: #eef4ff;
        border: 1.5px solid #3a7bd5;
        border-left: 4px solid #3a7bd5;
        border-radius: 6px;
        padding: .7rem 1rem;
        color: #1a3a6e;
        font-size: .84rem;
    }

    .box-ok {
        background: #f0fff5;
        border: 1.5px solid #28a745;
        border-radius: 6px;
        padding: .5rem .9rem;
        color: #1a5c2e;
        font-size: .84rem;
    }

    .box-warn {
        background: #fffbea;
        border: 1.5px solid #d97706;
        border-left: 4px solid #d97706;
        border-radius: 6px;
        padding: .65rem 1rem;
        color: #7c4700;
        font-size: .84rem;
    }

    .box-cf {
        background: #fff2f2;
        border: 1.5px solid #d63030;
        border-left: 4px solid #d63030;
        border-radius: 6px;
        padding: .7rem 1rem;
        margin: .4rem 0;
    }

    .cf-lbl {
        font-family: 'IBM Plex Mono', monospace;
        font-size: .66rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: .1em;
        color: #d63030;
        margin-bottom: .3rem;
    }

    /* <<<<<<<<<<<<<<<<<<<< TABLAS >>>>>>>>>>>>>>>>>>>> */
            
    .tt-wrap {
        background: #fff;
        border: 1px solid #ddd;
        border-radius: 8px;
        overflow: hidden;
        margin: .8rem 0;
    }

    .tt-hdr {
        background: #1c1c1c;
        color: #f0ede8;
        padding: .5rem 1rem;
        font-family: 'IBM Plex Mono', monospace;
        font-size: .68rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: .12em;
    }

    /* <<<<<<<<<<<<<<<<<<<< ETIQUETAS >>>>>>>>>>>>>>>>>>>> */
            
    .b-es {
        display: inline-block;
        padding: 1px 7px;
        border-radius: 3px;
        font-family: 'IBM Plex Mono', monospace;
        font-size: .64rem;
        font-weight: 700;
        background: #e6f4e6;
        color: #1a5c1a;
        border: 1px solid #a8d8a8;
    }

    .b-en {
        display: inline-block;
        padding: 1px 7px;
        border-radius: 3px;
        font-family: 'IBM Plex Mono', monospace;
        font-size: .64rem;
        font-weight: 700;
        background: #e6eeff;
        color: #1a2e6e;
        border: 1px solid #a8beef;
    }
            
    .b-ef {
        display: inline-block;
        padding: 1px 7px;
        border-radius: 3px;
        font-family: 'IBM Plex Mono', monospace;
        font-size: .64rem;
        font-weight: 700;
        background: #fff8e6;
        color: #7a4e00;
        border: 1px solid #f5d580;
    }

    .b-track {
        display: inline-block;
        padding: 1px 7px;
        border-radius: 3px;
        font-family: 'IBM Plex Mono', monospace;
        font-size: .62rem;
        font-weight: 700;
        background: #f3e8ff;
        color: #5b21b6;
        border: 1px solid #c4b5fd;
    }

    /* <<<<<<<<<<<<<<<<<<<< FORMULARIOS >>>>>>>>>>>>>>>>>>>> */
            
    .stSelectbox label,
    .stMultiSelect label,
    .stCheckbox label,
    .stTextInput label {
        color: #1a1a1a !important;
    }

    .stTextInput input {
        color: #1a1a1a !important;
        background: #fff !important;
    }

    /* <<<<<<<<<<<<<<<<<<<< DATAFRAME >>>>>>>>>>>>>>>>>>>> */
            
    div[data-testid="stDataFrame"] {
        border-radius: 6px;
        overflow: hidden;
        border: 1px solid #ddd;
    }
    </style>
""", unsafe_allow_html=True)

###########################################################
# Aquí están los códigos de todas las asignaturas hasta el
# POD de 2026. Sin embargo, como pasa con las 
# intensificaciones, es posible extraer nuevas asignaturas
# del POD y, con el código que se incluye en el excel de 
# los idiomas, cargar el código correspondiente.
###########################################################

TABLA_CODIGOS = {
    "TC":"42303","FGE":"42304","FP I":"42302","FFI":"42301","CMN":"42300",
    "SI":"42309","EC":"42307","RC I":"42308","FP II":"42306","AMD":"42305",
    "SSOO I":"42313","LOG":"42310","OC":"42311","EEDD":"42312","IS I":"42314",
    "BBDD":"42319","EST":"42315","RC II":"42318","PCTR":"42317","MP":"42316",
    "SSDD":"42322","AC":"42323","SSII":"42321","IPO I":"42320","IS II":"42324",
    "API":"42325",
    "TAC":"42342","DA":"42344","PD":"42343","SBC":"42345",
    "SIE":"42329","DS":"42327","IR":"42326","DBD":"42328",
    "DSBM":"42335","SSOO II":"42334","GAR":"42336","DIR":"42337",
    "IPO II":"42351","DGR":"42352","ISI":"42350","GSI":"42353",
    "PIS":"42330","SSS":"42333","GPS":"42332","CSW":"42331",
    "SR":"42340","CA":"42338","SSEE":"42339","PISS":"42341",
    "MM":"42356","SSI":"42357","TSW":"42354","CE":"42355",
    "DSI":"42347","MD":"42348","PL":"42349","SMA":"42346",
    "IWS":"42367","ASI":"42366","VARP":"42362",
    "AG":"42380","CRIP":"42364","DRI":"42371",
    "IAV":"42378","DGA":"42377","RA":"42361","VRV":"42379",
}

###########################################################
# También he añadido los nombres de las diversas
# asignaturas que hay en el grado, para poder visualizar el
# nombre de la asignatura en lugar de únicamente las
# siglas.
###########################################################

NOMBRES_ASIG = {
    "TC":"Tecnología de Computadores",
    "FGE":"Fundamentos de Gestión Empresarial",
    "FP I":"Fundamentos de Programación I",
    "FFI":"Fundamentos Físicos de la Informática",
    "CMN":"Cálculo y Métodos Numéricos",
    "SI":"Sistemas de Información",
    "EC":"Estructuras de Computadores",
    "RC I":"Redes de Computadores I",
    "FP II":"Fundamentos de Programación II",
    "AMD":"Álgebra y Matemática Discreta",
    "SSOO I":"Sistemas Operativos I",
    "LOG":"Lógica","OC":"Organización de Computadores",
    "EEDD":"Estructura de Datos",
    "IS I":"Ingeniería del Software I",
    "BBDD":"Bases de Datos",
    "EST":"Estadística",
    "RC II":"Redes de Computadores II",
    "PCTR":"Programación Concurrente y en Tiempo Real",
    "MP":"Metodología de la Programación",
    "SSDD":"Sistemas Distribuidos",
    "AC":"Arquitectura de Computadores",
    "SSII":"Sistemas Inteligentes",
    "IPO I":"Interacción Persona-Ordenador I",
    "IS II":"Ingeniería del Software II",
    "API":"Aspectos Profesionales de la Informática",
    "TAC":"Teoría de Autómatas y Computación",
    "DA":"Diseño de Algoritmos",
    "PD":"Programación Declarativa",
    "SBC":"Sistemas Basados en el Conocimiento",
    "SIE":"Sistemas de Información Empresariales",
    "DS":"Diseño de Software",
    "IR":"Ingeniería de Requisitos",
    "DBD":"Desarrollo de Bases de Datos",
    "DSBM":"Diseño de Sistemas Basados en Microprocesador",
    "SSOO II":"Sistemas Operativos II",
    "GAR":"Gestión y Administración de Redes",
    "DIR":"Diseño de Infraestructura Red",
    "IPO II":"Interacción Persona-Ordenador II",
    "DGR":"Diseño y Gestión de Redes",
    "ISI":"Integración de Sistemas Informáticos",
    "GSI":"Gestión de Sistemas de Información",
    "PIS":"Procesos de Ingeniería del Software",
    "SSS":"Seguridad de Sistemas Software",
    "GPS":"Gestión de Proyectos Software",
    "CSW":"Calidad de Sistemas Software",
    "SR":"Seguridad en Redes",
    "CA":"Computadores Avanzados",
    "SSEE":"Sistemas Empotrados",
    "PISS":"Planificación e Integración de Sistemas y Servicios",
    "MM":"Multimedia",
    "SSI":"Seguridad de Sistemas Informáticos",
    "TSW":"Tecnología y Sistemas Web",
    "CE":"Comercio Electrónico",
    "DSI":"Diseño de Sistemas Interactivos",
    "MD":"Minería de Datos",
    "PL":"Procesadores de Lenguajes",
    "SMA":"Sistemas Multiagentes",
    "IWS":"Ingeniería Web y de Servicios",
    "ASI":"Auditoría de Sistemas de Información",
    "VARP":"Visión Artificial y Reconocimiento de Patrones",
    "AG":"Aceleradores Gráficos",
    "CRIP":"Criptografía",
    "DRI":"Dispositivos y Redes Inalámbricos",
    "IAV":"Inteligencia Artificial en Videojuegos",
    "DGA":"Diseño Gráfico y Animación",
    "RA":"Robótica Autónoma",
    "VRV":"Videojuegos y Realidad Virtual",
}

IDIOMAS_DEFAULT = {
    "TC":{"Castellano":True,"Inglés":False,"EF":False},
    "FGE":{"Castellano":True,"Inglés":True,"EF":False},
    "FP I":{"Castellano":True,"Inglés":True,"EF":False},
    "FFI":{"Castellano":True,"Inglés":True,"EF":False},
    "CMN":{"Castellano":True,"Inglés":True,"EF":False},
    "SI":{"Castellano":True,"Inglés":True,"EF":False},
    "EC":{"Castellano":True,"Inglés":True,"EF":False},
    "RC I":{"Castellano":True,"Inglés":True,"EF":False},
    "FP II":{"Castellano":True,"Inglés":True,"EF":False},
    "AMD":{"Castellano":True,"Inglés":False,"EF":False},
    "SSOO I":{"Castellano":True,"Inglés":True,"EF":False},
    "LOG":{"Castellano":True,"Inglés":False,"EF":False},
    "OC":{"Castellano":True,"Inglés":True,"EF":False},
    "EEDD":{"Castellano":True,"Inglés":True,"EF":False},
    "IS I":{"Castellano":True,"Inglés":True,"EF":False},
    "BBDD":{"Castellano":True,"Inglés":False,"EF":True},
    "EST":{"Castellano":True,"Inglés":True,"EF":False},
    "RC II":{"Castellano":True,"Inglés":False,"EF":False},
    "PCTR":{"Castellano":True,"Inglés":True,"EF":False},
    "MP":{"Castellano":True,"Inglés":True,"EF":False},
    "SSDD":{"Castellano":True,"Inglés":False,"EF":False},
    "AC":{"Castellano":True,"Inglés":True,"EF":False},
    "SSII":{"Castellano":True,"Inglés":True,"EF":False},
    "IPO I":{"Castellano":True,"Inglés":True,"EF":False},
    "IS II":{"Castellano":True,"Inglés":True,"EF":False},
    "API":{"Castellano":True,"Inglés":False,"EF":False},
    "TAC":{"Castellano":True,"Inglés":False,"EF":True},
    "DA":{"Castellano":True,"Inglés":False,"EF":False},
    "PD":{"Castellano":True,"Inglés":False,"EF":False},
    "SBC":{"Castellano":True,"Inglés":False,"EF":False},
    "SIE":{"Castellano":True,"Inglés":False,"EF":False},
    "DS":{"Castellano":True,"Inglés":False,"EF":False},
    "IR":{"Castellano":True,"Inglés":False,"EF":False},
    "DBD":{"Castellano":True,"Inglés":False,"EF":False},
    "DSBM":{"Castellano":True,"Inglés":False,"EF":False},
    "SSOO II":{"Castellano":True,"Inglés":False,"EF":True},
    "GAR":{"Castellano":True,"Inglés":False,"EF":False},
    "DIR":{"Castellano":True,"Inglés":False,"EF":False},
    "IPO II":{"Castellano":True,"Inglés":False,"EF":False},
    "DGR":{"Castellano":True,"Inglés":False,"EF":False},
    "ISI":{"Castellano":True,"Inglés":False,"EF":False},
    "GSI":{"Castellano":True,"Inglés":False,"EF":False},
    "PIS":{"Castellano":True,"Inglés":False,"EF":False},
    "SSS":{"Castellano":True,"Inglés":False,"EF":False},
    "GPS":{"Castellano":True,"Inglés":False,"EF":True},
    "CSW":{"Castellano":True,"Inglés":False,"EF":True},
    "SR":{"Castellano":True,"Inglés":False,"EF":True},
    "CA":{"Castellano":True,"Inglés":False,"EF":False},
    "SSEE":{"Castellano":True,"Inglés":False,"EF":True},
    "PISS":{"Castellano":True,"Inglés":False,"EF":True},
    "MM":{"Castellano":True,"Inglés":False,"EF":True},
    "SSI":{"Castellano":True,"Inglés":False,"EF":True},
    "TSW":{"Castellano":True,"Inglés":False,"EF":False},
    "CE":{"Castellano":True,"Inglés":False,"EF":False},
    "MD":{"Castellano":True,"Inglés":False,"EF":True},
    "PL":{"Castellano":True,"Inglés":False,"EF":True},
    "SMA":{"Castellano":True,"Inglés":False,"EF":False},
    "DSI":{"Castellano":True,"Inglés":False,"EF":False},
    "IWS":{"Castellano":False,"Inglés":True,"EF":False},
    "ASI":{"Castellano":True,"Inglés":False,"EF":False},
    "VARP":{"Castellano":True,"Inglés":False,"EF":False},
    "AG":{"Castellano":True,"Inglés":False,"EF":False},
    "CRIP":{"Castellano":True,"Inglés":False,"EF":True},
    "DRI":{"Castellano":True,"Inglés":False,"EF":True},
    "IAV":{"Castellano":False,"Inglés":True,"EF":False},
    "DGA":{"Castellano":False,"Inglés":True,"EF":False},
    "RA":{"Castellano":False,"Inglés":True,"EF":False},
    "VRV":{"Castellano":False,"Inglés":True,"EF":False},
}

TRADUCCION_POR_CODIGO = {
    42300:{"en":"Calculus and Numerical Methods","es":"Cálculo y Métodos Numéricos"},
    42301:{"en":"Physical Foundations of Computing","es":"Fundamentos Físicos de la Informática"},
    42302:{"en":"Programming Fundamentals I","es":"Fundamentos de Programación I"},
    42303:{"en":"Computer Technology","es":"Tecnología de Computadores"},
    42304:{"en":"Fundamentals of Business Management","es":"Fundamentos de Gestión Empresarial"},
    42305:{"en":"Algebra and Discrete Mathematics","es":"Álgebra y Matemática Discreta"},
    42306:{"en":"Programming Fundamentals II","es":"Fundamentos de Programación II"},
    42307:{"en":"Computer Structure","es":"Estructuras de Computadores"},
    42308:{"en":"Computer Networks I","es":"Redes de Computadores I"},
    42309:{"en":"Information Systems","es":"Sistemas de Información"},
    42310:{"en":"Logic","es":"Lógica"},
    42311:{"en":"Computer Organization","es":"Organización de Computadores"},
    42312:{"en":"Data Structures","es":"Estructura de Datos"},
    42313:{"en":"Operating Systems I","es":"Sistemas Operativos I"},
    42314:{"en":"Software Engineering I","es":"Ingeniería del Software I"},
    42315:{"en":"Statistics","es":"Estadística"},
    42316:{"en":"Programming Methodology","es":"Metodología de la Programación"},
    42317:{"en":"Concurrent and Real-Time Programming","es":"Programación Concurrente y en Tiempo Real"},
    42318:{"en":"Computer Networks II","es":"Redes de Computadores II"},
    42319:{"en":"Databases","es":"Bases de Datos"},
    42320:{"en":"Human-Computer Interaction I","es":"Interacción Persona-Ordenador I"},
    42321:{"en":"Intelligent Systems","es":"Sistemas Inteligentes"},
    42322:{"en":"Distributed Systems","es":"Sistemas Distribuidos"},
    42323:{"en":"Computer Architecture","es":"Arquitectura de Computadores"},
    42324:{"en":"Software Engineering II","es":"Ingeniería del Software II"},
    42325:{"en":"Professional Aspects of Computing","es":"Aspectos Profesionales de la Informática"},
    42326:{"en":"Requirements Engineering","es":"Ingeniería de Requisitos"},
    42327:{"en":"Software Design","es":"Diseño de Software"},
    42328:{"en":"Database Development","es":"Desarrollo de Bases de Datos"},
    42329:{"en":"Enterprise Information Systems","es":"Sistemas de Información Empresariales"},
    42330:{"en":"Software Engineering Processes","es":"Procesos de Ingeniería del Software"},
    42331:{"en":"Software Systems Quality","es":"Calidad de Sistemas Software"},
    42332:{"en":"Software Project Management","es":"Gestión de Proyectos Software"},
    42333:{"en":"Software Security","es":"Seguridad de Sistemas Software"},
    42334:{"en":"Operating Systems II","es":"Sistemas Operativos II"},
    42335:{"en":"Microprocessor-Based Systems Design","es":"Diseño de Sistemas Basados en Microprocesador"},
    42336:{"en":"Network Management and Administration","es":"Gestión y Administración de Redes"},
    42337:{"en":"Network Infrastructure Design","es":"Diseño de Infraestructura Red"},
    42338:{"en":"Advanced Computers","es":"Computadores Avanzados"},
    42339:{"en":"Embedded Systems","es":"Sistemas Empotrados"},
    42340:{"en":"Network Security","es":"Seguridad en Redes"},
    42341:{"en":"Systems Planning and Integration","es":"Planificación e Integración de Sistemas y Servicios"},
    42342:{"en":"Automata Theory and Computation","es":"Teoría de Autómatas y Computación"},
    42343:{"en":"Declarative Programming","es":"Programación Declarativa"},
    42344:{"en":"Algorithm Design","es":"Diseño de Algoritmos"},
    42345:{"en":"Knowledge-Based Systems","es":"Sistemas Basados en el Conocimiento"},
    42346:{"en":"Multi-Agent Systems","es":"Sistemas Multiagentes"},
    42347:{"en":"Interactive Systems Design","es":"Diseño de Sistemas Interactivos"},
    42348:{"en":"Data Mining","es":"Minería de Datos"},
    42349:{"en":"Language Processors","es":"Procesadores de Lenguajes"},
    42350:{"en":"Computer Systems Integration","es":"Integración de Sistemas Informáticos"},
    42351:{"en":"Human-Computer Interaction II","es":"Interacción Persona-Ordenador II"},
    42352:{"en":"Network Design and Management","es":"Diseño y Gestión de Redes"},
    42353:{"en":"Information Systems Management","es":"Gestión de Sistemas de Información"},
    42354:{"en":"Web Technologies and Systems","es":"Tecnología y Sistemas Web"},
    42356:{"en":"Multimedia","es":"Multimedia"},
    42357:{"en":"Computer Systems Security","es":"Seguridad de Sistemas Informáticos"},
    42355:{"en":"E-Commerce","es":"Comercio Electrónico"},
    42361:{"en":"Autonomous Robotics","es":"Robótica Autónoma"},
    42362:{"en":"Computer Vision and Pattern Recognition","es":"Visión Artificial y Reconocimiento de Patrones"},
    42364:{"en":"Cryptography","es":"Criptografía"},
    42366:{"en":"Information Systems Auditing","es":"Auditoría de Sistemas de Información"},
    42367:{"en":"Web and Service Engineering","es":"Ingeniería Web y de Servicios"},
    42371:{"en":"Wireless Devices and Networks","es":"Dispositivos y Redes Inalámbricos"},
    42377:{"en":"Graphic Design and Animation","es":"Diseño Gráfico y Animación"},
    42378:{"en":"Artificial Intelligence in Video Games","es":"Inteligencia Artificial en Videojuegos"},
    42379:{"en":"Video Games and Virtual Reality","es":"Videojuegos y Realidad Virtual"},
    42380:{"en":"Graphics Accelerators","es":"Aceleradores Gráficos"},
}

###########################################################
# Las intensificiaciones están hardcodeadas, lo que
# significa que las asignaturas están asignadas a la
# intensificación correspondiente. Sin embargo, es posible
# obtener más intensificaciones si se detectan en el POD.
###########################################################

INTENSIFICACIONES = {
    42342:"CO", 42343:"CO", 42344:"CO", 42345:"CO", 42346:"CO", 42347:"CO", 42348:"CO", 42349:"CO", #Intesificación de Gilberth
    42334:"IC", 42335:"IC", 42336:"IC", 42337:"IC", 42338:"IC", 42339:"IC", 42340:"IC", 42341:"IC", #Intensificación de Diego
    42326:"IS", 42327:"IS", 42328:"IS", 42329:"IS", 42330:"IS", 42331:"IS", 42332:"IS", 42333:"IS", #Intensificación de Julio y mía
    42350:"TI", 42351:"TI", 42352:"TI", 42353:"TI", 42354:"TI", 42355:"TI", 42356:"TI", 42357:"TI"  #Intensificación de Paco y Elías
}

ABREVIATURAS_INTESIFICACIONES: set = set()

PATRONESOPT = [r'O[123]', r'Optativas?', r'Optativas? O', r'Menciones?']

NOMBRES_INTENSIFICACIONES = {
    "CO": "Computación (CO)",
    "IC": "Ing. de Computadores (IC)",
    "IS": "Ing. del Software (IS)",
    "TI": "Tecnologías de la Información (TI)",
}

###########################################################
# Aquí he añadido las configuraciones básicas de la página
# del TFG usando streamlit, donde configuro el título de la
# página (la pestaña), la disposición de la página (puse
# wide para que ocupara toda la pantalla disponible) y el
# estado inicial de la barra lateral.
###########################################################

st.set_page_config(
    page_title="Gestor de horario", 
    layout="wide",
    initial_sidebar_state="expanded",
)

###########################################################
# A continuación, se incluye la configuración de días y de
# colores para la tabla del horario de los estudiantes,
# generada a partir de la selección de las asignaturas y
# el grupo de teoría y prácticas correspondiente.
###########################################################

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
DIASTEORIAAPI = {"Lunes", "Martes"}
DIASLABAPI = {"Miércoles", "Jueves"}
DIAS_ES = ["Lunes","Martes","Miércoles","Jueves","Viernes"]
DIAS_EN = ["Monday","Tuesday","Wednesday","Thursday","Friday"]

COLORES = [
    ("#d1fae5", "#065f46"),
    ("#dbeafe", "#1d4ed8"),
    ("#dcfce7", "#166534"),
    ("#e0f2fe", "#0c4a6e"),
    ("#ede9fe", "#5b21b6"),
    ("#f0fdf4", "#14532d"),
    ("#f8fafc", "#0f172a"),
    ("#fce7f3", "#9d174d"),
    ("#fdf4ff", "#6b21a8"),
    ("#fee2e2", "#991b1b"),
    ("#fef3c7", "#92400e"),
    ("#fff7ed", "#7c2d12"),
]

###########################################################
# Esta paleta de colores sirve para los diferentes colores
# que conforman las etiquetas de las asignaturas en el
# listado de asignaturas.
###########################################################

PALETA = [
    "D1FAE5", 
    "DBEAFE", 
    "DCFCE7", 
    "E0F2FE", 
    "EDE9FE",
    "FCE7F3", 
    "FDF4FF", 
    "FEE2E2", 
    "FEF3C7", 
    "FFF7ED",
]

###########################################################
# En esta parte, se incluye la traducción de los distintos
# textos que se encuentran en la página, de forma que, si
# se cambia de idioma a inglés o a español, se seleccione
# el texto correcto.
###########################################################

UI = {
    "es":{
        "acceso":"Acceso protegido por contraseña.",
        "add":"Añadir",
        "administrar_datos":"Administración de datos del POD. Solo para coordinadores.",
        "anio":"Curso",
        "api_lab":"Franja de prácticas (API)",
        "api_teoria":"Franja de teoría (API)",
        "asig_disp":"asignaturas disponibles",
        "asigs":"asig.",
        "bil":"Bilingüe",
        "borrar_todo":"Limpiar horario",
        "cambios_guardados":"Cambios guardados.",
        "cargando":"Procesando...",
        "cargar_idiomas":"Excel de idiomas (.xlsx)",
        "cargar_pod":"Cargar POD / Excel",
        "consejo_solape":"Cambia el grupo o continúa de todas formas.",
        "copiar_al_portapapeles":"Copiar al portapapeles",
        "copiar_horario":"Copiar horario al portapapeles",
        "copy_ok":"¡Copiado!",
        "cuatri":"Cuatrimestre",
        "cuatri1":"1.º Cuatrimestre",
        "cuatri1_etiqueta":"1.er Cuatrimestre",
        "cuatri2":"2.º Cuatrimestre",
        "cuatri2_etiqueta":"2.º Cuatrimestre",
        "desbloquear":"Desbloquear",
        "descargar_horario":"Descargar horario (tabla)",
        "descargar_horario_excel":"Descargar horario en Excel",
        "descargar_idiomas":"Descargar idiomas actuales",
        "ds_export_btn":"Descargar Excel completo",
        "editar_tabla":"Ver / Editar Excel",
        "ef":"English-Friendly",
        "eliminar":"Quitar",
        "error_carga":"Error al procesar el archivo",
        "error_idiomas":"Error al procesar el Excel de idiomas.",
        "etiqueta_carga":"PDF del POD o Excel (.pdf / .xlsx)",
        "etiqueta_horario":"Horario",
        "excel_cargado":"Excel cargado correctamente.",
        "exportar_datos":"Exportar tabla / datos",
        "exportar_tabla":"Exportar Excel",
        "filas":"filas",
        "fuente_de_datos":"Fuente de Datos",
        "generar_horario":"Crear Horario",
        "guardar_cambios":"Guardar cambios",
        "grp":"Grupo",
        "grp_aleatorio":"Cualquiera",
        "idioma":"Idioma",
        "idioma_ef":"🇬🇧 EF",
        "idioma_en":"🇬🇧 Inglés",
        "idioma_es":"🇪🇸 Castellano",
        "idiomas_cargados":"Idiomas actualizados correctamente.",
        "lab":"Práctica",
        "leyenda":"Leyenda",
        "no_datos":"Carga datos en Fuente de Datos primero.",
        "no_lab":"Ninguna asignatura tiene dos sesiones de prácticas.",
        "no_solapes":"Sin solapamientos en la selección actual.",
        "omitir_solapes":"Continuar con solapamientos",
        "ordenar":"Clic en encabezado para ordenar",
        "pista":"Selecciona asignaturas y elige grupo e idioma",
        "pista_idioma":"Carga el Excel con columnas: Código Grado, Grado, Código Asignatura, Nombre Asignatura, Español, English Friendly, Inglés. Solo se cargarán las filas con Código Grado = 406.",
        "pista_lab":"Elige qué sesión de prácticas quieres cursar (cuando hay dos).",
        "pod_cargado":"Excel generado correctamente desde el PDF.",
        "pswd":"Contraseña",
        "pswd_incorrecta":"Contraseña incorrecta.",
        "seleccion":"asignatura(s) seleccionada(s)",
        "sesion":"Sesión de prácticas",
        "sesion1":"Sesión 1",
        "sesion2":"Sesión 2",
        "sin_datos":"No hay datos cargados todavía.",
        "solape":"coinciden en",
        "subtitulo":"Grado en Ingeniería Informática · UCLM",
        "tabla_idiomas":"Cargar Idiomas",
        "teoria":"Teoría",
        "texto_busqueda":"Buscar por código o nombre... (Intro para añadir)",        
        "titulo":"Generador de Horario",
        "titulo_admin":"Fuente de Datos",
        "track_blocked":"Ya tienes asignaturas de {track}. Esta asignatura pertenece a {other}.",
        "titulo_horario":"Crear Horario",
        "titulo_solapes":"Solapamiento",
        "track_warn":"Solo puedes elegir asignaturas de una intensificación.",
    },
    "en":{
        "acceso":"Password-protected access.",
        "add":"Add",
        "administrar_datos":"POD data management. For coordinators only.",
        "anio":"Year",
        "api_lab":"Lab slot (API)",
        "api_teoria":"Theory slot (API)",
        "asig_disp":"subjects available",
        "asigs":"subj.",
        "bil":"Bilingual",
        "borrar_todo":"Clear schedule",
        "cambios_guardados":"Changes saved.",
        "cargando":"Processing...",
        "cargar_idiomas":"Language Excel file (.xlsx)",
        "cargar_pod":"Load POD / Excel",
        "consejo_solape":"Change the group or continue anyway.",
        "copiar_al_portapapeles":"Copy to clipboard",        
        "copiar_horario":"Copy schedule to clipboard",
        "copy_ok":"Copied!",
        "cuatri":"Semester",
        "cuatri1":"1st Semester",
        "cuatri1_etiqueta":"1st Semester",
        "cuatri2":"2nd Semester",
        "cuatri2_etiqueta":"2nd Semester",
        "desbloquear":"Unlock",
        "descargar_horario":"Download schedule (table)",
        "descargar_horario_excel":"Download schedule as Excel",
        "descargar_idiomas":"Download current languages",
        "ds_export_btn":"Download full Excel",
        "editar_tabla":"View / Edit Excel",
        "ef":"English-Friendly",
        "eliminar":"Remove",
        "error_carga":"Error processing file",
        "error_idiomas":"Error processing language Excel.",
        "etiqueta_carga":"POD PDF or Excel (.pdf / .xlsx)",
        "etiqueta_horario":"Timetable",
        "excel_cargado":"Excel loaded successfully.",
        "exportar_datos":"Export table / data",
        "exportar_tabla":"Export Excel",
        "filas":"rows",
        "fuente_de_datos":"Data Source",
        "generar_horario":"Build Schedule",
        "guardar_cambios":"Save changes",
        "grp":"Group",
        "grp_aleatorio":"Any",
        "idioma":"Language",
        "idioma_ef":"🇬🇧 EF",
        "idioma_en":"🇬🇧 English",
        "idioma_es":"🇪🇸 Spanish",
        "idiomas_cargados":"Languages updated successfully.",
        "lab":"Lab",
        "leyenda":"Legend",        
        "no_datos":"Load data in Data Source first.",
        "no_lab":"No subject has two lab sessions.",
        "no_solapes":"No overlaps in the current selection.",
        "omitir_solapes":"Continue with overlaps",
        "ordenar":"Click a column header to sort",
        "pista":"Select subjects and choose group and language",
        "pista_idioma":"Load an Excel with columns: Código Grado, Grado, Código Asignatura, Nombre Asignatura, Español, English Friendly, Inglés. Only rows with Código Grado = 406 will be imported.",
        "pista_lab":"Choose which lab session to attend (when two are available).",
        "pod_cargado":"Excel generated successfully from the PDF.",
        "pswd":"Password",
        "pswd_incorrecta":"Incorrect password.",
        "seleccion":"subject(s) selected",
        "sesion":"Lab session",
        "sesion1":"Session 1",
        "sesion2":"Session 2",
        "sin_datos":"No data loaded yet.",
        "solape":"overlap at",
        "subtitulo":"Computer Science Engineering · UCLM",
        "tabla_idiomas":"Load Languages",
        "teoria":"Lecture",
        "texto_busqueda":"Search by code or name... (Enter to add)",
        "titulo":"Schedule Builder",
        "titulo_admin":"Data Source",
        "titulo_horario":"Build Schedule",
        "titulo_solapes":"Overlap detected",
        "track_blocked":"You already have subjects from {track}. This subject belongs to {other}.",
        "track_warn":"You can only choose subjects from one specialisation track.",
    },
}


###########################################################
# Estas variables que se presentan a continuación sirven
# para mantener los cambios entre diferentes sesiones
# una vez desplegado el servidor. En modo local no
# funciona, pues se reinician los cambios cada vez y no hay
# almacenamiento.
###########################################################

DIRECTORIO = os.path.join(os.path.dirname(__file__) if "__file__" in dir() else ".", ".pod_data")
ASIGNATURAS = os.path.join(DIRECTORIO, "asignaturas.parquet")
IDIOMAS = os.path.join(DIRECTORIO, "lang_overrides.json")

PSWD = "horario2025"

def comprobar_directorio():
    os.makedirs(DIRECTORIO, exist_ok=True)

def guardar_cambios(asig, cambiaIdiomas):
    try:
        comprobar_directorio()
        if asig is not None and not asig.empty: asig.to_parquet(ASIGNATURAS, index=False)
        with open(IDIOMAS, "w", encoding="utf-8") as f: json.dump({str(k): v for k, v in cambiaIdiomas.items()}, f, ensure_ascii=False)
    except Exception:
        pass

def cargar_cambios():
    asignaturas = None
    idiomas = {}

    try:
        if os.path.exists(ASIGNATURAS): asignaturas = pandas.read_parquet(ASIGNATURAS)
    except Exception:
        pass

    try:
        if os.path.exists(IDIOMAS):
            with open(IDIOMAS, "r", encoding="utf-8") as f: datos = json.load(f)
            idiomas = {int(k): v for k, v in datos.items()}
    except Exception:
        pass

    return asignaturas, idiomas

###########################################################
# En esta parte, me encargo de gestión el estado de la
# sesión, de forma que no se pierdan los cambios en las
# sesiones online. Además, también gestiono qué aparece al
# entrar a la página (la sección, idioma, erasmus, etc.).
###########################################################

if "app_initialized" not in st.session_state:
    datos_materias, datos_idiomas = cargar_cambios()
    st.session_state["df_asignaturas"] = datos_materias
    st.session_state["lang_overrides"] = datos_idiomas
    st.session_state["app_initialized"] = True

VALOR_DEFECTO = {
    "idioma": "es",
    "seccion": "horario",
    "modo_admin": False,
    "asignaturas_seleccionadas": {},
    "idiomas_seleccionados": {},
    "sesiones_practicas": {},
    "erasmus": False,
    "grupo_teoria_api": 1,
    "grupo_lab_api": 1,
    "indice_api_lab": 0,
    "horario_api_lab": None,
}
for dato, valor in VALOR_DEFECTO.items():
    if dato not in st.session_state:
        st.session_state[dato] = valor

###########################################################
# Aquí he declarado las funciones getters y el setter que 
# emplearé más adelante para hacer funcionar la página.
###########################################################

def get_infoAsignatura(coda):
    codAsignatura = convertirAEntero(coda)
    return TRADUCCION_POR_CODIGO.get(codAsignatura) if codAsignatura else None

def get_dias():
    return DIAS_EN if st.session_state.idioma == "en" else DIAS_ES

def get_intensificacion(codigo):
    k = convertirAEntero(codigo)
    return INTENSIFICACIONES.get(k) if k else None

def get_intensificacionPorFila(fila):
    valor = str(fila.get("Intensificacion", "") or "").strip()
    if valor:
        return valor
    return get_intensificacion(fila.get("Código"))

def get_huecos(fila, sesionLab=None):
    huecos = []
    for i in (1, 2):
        dia, hora = normalizar(fila.get(f"T{i}-día")), normalizar(fila.get(f"T{i}-hora"))
        if dia and hora:
            huecos.append(("T", dia, hora))
    for i in (1, 2):
        if sesionLab is not None and sesionLab != i:
            continue
        dia, hora = normalizar(fila.get(f"L{i}-día")), normalizar(fila.get(f"L{i}-hora"))
        if dia and hora:
            huecos.append(("L", dia, hora))
    return huecos

def get_tablas(tamanio, paginas):
    listaIdentificadores = {}
    with pdfplumber.open(io.BytesIO(tamanio)) as pdf:
        for np in paginas:
            pagina = pdf.pages[np - 1]
            for table in pagina.extract_tables():
                hdr = table[0][0] or ""
                m = re.search(r"\(([^)]+)\)", hdr)
                aula_raw = m.group(1).strip() if m else ""
                if "/" in aula_raw:
                    parts = [p.strip() for p in aula_raw.split("/", 1)]
                    lab_parts = [p for p in parts if "lab" in p.lower()]
                    room_parts = [p for p in parts if any(c.isdigit() for c in p)]
                    aula_lab = lab_parts[0] if lab_parts else parts[1]
                    aula_teoria = room_parts[0] if room_parts else parts[0]
                else:
                    if "lab" in aula_raw.lower():
                        aula_lab = aula_raw
                        aula_teoria = ""
                    elif any(c.isdigit() for c in aula_raw):
                        aula_teoria = aula_raw
                        aula_lab = "lab"
                    else:
                        aula_teoria = aula_raw
                        aula_lab = "lab"
                identificador = hdr.split("(")[0].strip()
                
                dataframe = pandas.DataFrame(table[1:], columns=table[0])
                dataframe["Identificador"] = identificador
                dataframe["AulaTeoria"] = aula_teoria
                dataframe["AulaLab"] = aula_lab
                dataframe.rename(columns={dataframe.columns[0]: "Hora"}, inplace=True)
                dataframe["Hora"] = dataframe["Hora"].str.replace("\n\n", "-")

                if identificador in listaIdentificadores:
                    listaIdentificadores[identificador] = pandas.concat([listaIdentificadores[identificador], dataframe])
                else:
                    listaIdentificadores[identificador] = dataframe
    return pandas.concat(list(listaIdentificadores.values())).reset_index(drop=True)

def set_celda(cuatri, dia, hora, fondo, texto, disp, aula, icon, tipo):
    au = (f'<div style="font-size:.61rem;color:#555;margin-top:1px">🖈 {aula}</div>' if aula else "")
    cuatri[(dia, hora)] = cuatri.get((dia, hora), "") + (
        f'<div style="background:{fondo};border-left:3px solid {texto};border-radius:4px;' +
        f'padding:4px 7px;font-size:.74rem;font-weight:600;margin:2px 0;color:{texto}">' +
        f'{disp}{au}<div style="font-weight:400;font-size:.65rem;opacity:.85">{icon} {tipo}</div></div>'
    )

###########################################################
# Aquí he declarado las funciones auxiliares que emplearé
# más adelante para hacer funcionar la página.
###########################################################

def obtenerTexto(x): 
    return UI[st.session_state.idioma].get(x, x)

def normalizar(x):
    s = str(x or "").strip()
    return "" if s.lower() == "nan" else s

def convertirAEntero(x):
    try:
        return int(str(x))
    except Exception:
        return None

def traducirDia(dia):
    if st.session_state.idioma == "en":
        return dict(zip(DIAS_ES, DIAS_EN)).get(dia, dia)
    return dia

def convertirAMinutos(hora):
    try:
        particiones = str(hora).strip().split(":")
        return int(particiones[0]) * 60 + int(particiones[1])
    except Exception:
        return 9999

def mostrarNombre(abrev, codigo=None):
    info = get_infoAsignatura(codigo) if codigo else None
    if info:
        return info["en"] if st.session_state.idioma == "en" else info["es"]
    return abrev

def obtenerBanderasIdioma(abrev, codigo=None):
    sobreescritura = st.session_state.lang_overrides
    coda = convertirAEntero(codigo) if codigo else None
    if coda and coda in sobreescritura:
        dia = sobreescritura[coda]
        espaniol = bool(dia.get("Español", False))
        ingles = bool(dia.get("Inglés", False))
        ef = bool(dia.get("EF", False))
        bil = espaniol and ingles and not ef
        ingles = ingles and not espaniol
        return bil, ef, ingles
    dia = IDIOMAS_DEFAULT.get(abrev, {"Castellano": True, "Inglés": False, "EF": False})
    espaniol = bool(dia.get("Castellano", True))
    ingles = bool(dia.get("Inglés", False))
    ef = bool(dia.get("EF", False))
    bil = espaniol and ingles and not ef
    ingles = ingles and not espaniol
    return bil, ef, ingles

def primeraIntensificacion():
    dataframe = st.session_state.df_asignaturas
    if dataframe is None:
        return None
    for abrev in st.session_state.asignaturas_seleccionadas:
        filas = dataframe[dataframe["Asignatura"] == abrev]
        if filas.empty:
            continue
        primeraIntensificacion = get_intensificacionPorFila(filas.iloc[0])
        if primeraIntensificacion:
            return primeraIntensificacion
    return None

def comprobarSolapes(elecciones, dataframe, sesionLabElegida):
    huecos = {}
    for abrev, grupo in elecciones:
        filas = (dataframe[dataframe["Asignatura"] == abrev] if grupo == "auto"
                else dataframe[(dataframe["Asignatura"] == abrev) & (dataframe["Grupo"].astype(str) == str(grupo))])
        if filas.empty:
            continue
        fila = filas.iloc[0]
        cuatri = str(fila.get("Cuatrimestre", "")).strip()
        if not cuatri:
            cuatri = "1"
        labElegido = sesionLabElegida.get((abrev, grupo))
        for _, dia, hora in get_huecos(fila, labElegido):
            huecos.setdefault((dia, hora, cuatri), []).append((abrev, grupo))
    conflictos, detectados = [], set()
    for (dia, hora, _), asignaturasSeleccionadas in huecos.items():
        for i in range(len(asignaturasSeleccionadas)):
            for j in range(i + 1, len(asignaturasSeleccionadas)):
                tupla = tuple(sorted([asignaturasSeleccionadas[i], asignaturasSeleccionadas[j]]))
                if tupla not in detectados:
                    detectados.add(tupla)
                    ex = next((c for c in conflictos if (c[0], c[1]) == tupla), None)
                    if ex:
                        ex[2].append((dia, hora))
                    else:
                        conflictos.append([tupla[0], tupla[1], [(dia, hora)]])
    return conflictos

def aplicarCSSASeccion(texto):
    st.markdown(f"<div class='sec-hdr'>{texto}</div>", unsafe_allow_html=True)

def aplicarCSSACajaTexto(hora):
    st.markdown(f"<div class='box-info'>{hora}</div>", unsafe_allow_html=True)

def textoCorrecto(hora):
    st.markdown(f"<div class='box-ok'>{hora}</div>", unsafe_allow_html=True)

def advertencia(hora):
    st.markdown(f"<div class='box-warn'>⚠ {hora}</div>", unsafe_allow_html=True)

def textoComodin(titulo, cuerpo):
    st.markdown(f"<div class='box-cf'><div class='cf-lbl'>⚠ {titulo}</div>"
                f"<span style='font-size:.85rem;color:#1a1a1a'>{cuerpo}</span></div>",
                unsafe_allow_html=True)

def copiarTexto(dataframe):
    return dataframe.to_csv(sep="\t", index=False)

def botonCopiar(etiqueta, portapapeles, clave):
    b64 = base64.b64encode(portapapeles.encode("utf-8")).decode("ascii")
    todoCorrecto = "¡Copiado!" if st.session_state.idioma == "es" else "Copied!"
    html = f"""
    <style>

      #cb_btn_{clave} {{
        background:#1c1c1c;color:#f0ede8;border:none;border-radius:4px;
        font-family:'IBM Plex Mono',monospace;font-weight:600;font-size:.76rem;
        text-transform:uppercase;letter-spacing:.05em;padding:.42rem .9rem;
        cursor:pointer;
      }}
      #cb_ok_{clave} {{display:none;font-size:.8rem;color:#1a7e45;font-weight:600;margin-left:.5rem;}}
    
    </style>

    <button id="cb_btn_{clave}" onclick="(function(){{
      var texto = atob('{b64}');
      var long = new Uint8Array(texto.length);

      for(var i=0;i<texto.length;i++) long[i]=texto.charCodeAt(i);
      texto = new TextDecoder('utf-8').decode(long);
      navigator.clipboard.writeText(texto).then(function(){{
        var todoCorrecto = document.getElementById('cb_ok_{clave}');
        todoCorrecto.style.display='inline';
        setTimeout(function(){{todoCorrecto.style.display='none';}}, 2500);
      }}).catch(function(){{
        var ta = document.createElement('textarea');
        ta.value = texto; ta.style.position='fixed'; ta.style.opacity=0;
        document.body.appendChild(ta); ta.select();
        document.execCommand('copy');
        document.body.removeChild(ta);
        var todoCorrecto = document.getElementById('cb_ok_{clave}');
        todoCorrecto.style.display='inline';
        setTimeout(function(){{todoCorrecto.style.display='none';}}, 2500);
      }});
    }})()">🗐 {etiqueta}</button>
    <span id="cb_ok_{clave}">{todoCorrecto}</span>
    """
    componentes.html(html, height=45)

def crearTablaExcel(combinacionAsignaturas, sesionLabElegida, cambiaIdiomas):
    delgado = Side(style="thin", color="CCCCCC")
    border = Border(left=delgado, right=delgado, top=delgado, bottom=delgado)

    def agrupar(filtrarCuatri):
        huecos = {}
        for fila in combinacionAsignaturas:
            cuatri = str(fila.get("Cuatrimestre", "1")).strip()
            if cuatri != filtrarCuatri:
                continue
            abrev = fila.get("Asignatura", "")
            grupo = fila.get("Grupo", "")
            labElegido = sesionLabElegida.get((abrev, grupo))
            info = get_infoAsignatura(convertirAEntero(fila.get("Código")))
            nombre = info["es"] if info else abrev

            for i in (1, 2):
                dia = normalizar(fila.get(f"T{i}-día"))
                hora = normalizar(fila.get(f"T{i}-hora"))
                au = normalizar(fila.get(f"T{i}-aula", ""))
                if dia and hora:
                    textoCelda = f"{abrev}\nT G{grupo}"
                    if au:
                        textoCelda += f"\n🖈{au}"
                    huecos.setdefault((hora, dia), []).append(textoCelda)
            for i in (1, 2):
                if labElegido is not None and labElegido != i:
                    continue
                dia = normalizar(fila.get(f"L{i}-día"))
                hora = normalizar(fila.get(f"L{i}-hora"))
                au = normalizar(fila.get(f"L{i}-aula", "lab"))
                if dia and hora:
                    textoCelda = f"{abrev}\nL G{grupo}"
                    if au:
                        textoCelda += f"\n🖈{au}"
                    huecos.setdefault((hora, dia), []).append(textoCelda)
        return huecos

    wb = Workbook()
    wb.remove(wb.active)
    rellenarCabecera = PatternFill("solid", fgColor="1C1C1C")
    fuenteCabecera = Font(color="F0EDE8", bold=True, name="Courier New", size=9)

    def rellenarAbreviatura(abrev, listaAbreviaturas):
        indice = listaAbreviaturas.index(abrev) % len(PALETA) if abrev in listaAbreviaturas else 0
        return PatternFill("solid", fgColor=PALETA[indice])

    listaAbreviaturas = list({fila.get("Asignatura", "") for fila in combinacionAsignaturas})

    for cuatri, nombreHoja in [("1", "1er Cuatrimestre"), ("2", "2o Cuatrimestre")]:
        huecos = agrupar(cuatri)
        if not huecos:
            continue
        ws = wb.create_sheet(title=nombreHoja)
        horas = sorted(set(hora for (hora, _) in huecos), key=convertirAMinutos)
        if not horas:
            continue

        ws.column_dimensions["A"].width = 9
        ws.cell(1, 1, "Hora").fill = rellenarCabecera
        ws.cell(1, 1).font = fuenteCabecera
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(1, 1).border = border
        for indiceColumna, dia in enumerate(DIAS_ES, start=2):
            ws.column_dimensions[get_column_letter(indiceColumna)].width = 22
            c = ws.cell(1, indiceColumna, dia)
            c.fill = rellenarCabecera
            c.font = fuenteCabecera
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        for indiceFila, hora in enumerate(horas, start=2):
            ws.row_dimensions[indiceFila].height = 48
            hc = ws.cell(indiceFila, 1, hora)
            hc.fill = PatternFill("solid", fgColor="F5F4F0")
            hc.font = Font(name="Courier New", size=8, color="888888")
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border = border
            for indiceColumna, dia in enumerate(DIAS_ES, start=2):
                huecos_celda = huecos.get((hora, dia), [])
                txt = "\n".join(huecos_celda)
                abreviatura = huecos_celda[0].split("\n")[0] if huecos_celda else ""
                c = ws.cell(indiceFila, indiceColumna, txt)
                if abreviatura and abreviatura in listaAbreviaturas:
                    c.fill = rellenarAbreviatura(abreviatura, listaAbreviaturas)
                c.font = Font(name="Calibri", size=8)
                c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                c.border = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def crearExcelIdiomas(cambiaIdiomas):
    filas = []
    for coda, dia in cambiaIdiomas.items():
        info = get_infoAsignatura(coda)
        filas.append({
            "Código Grado": 406,
            "Código Asignatura": coda,
            "Nombre Asignatura": info["es"] if info else str(coda),
            "Español": "Si" if dia.get("Español") else "No",
            "English Friendly": "Si" if dia.get("EF") else "No",
            "Inglés": "Si" if dia.get("Inglés") else "No",
        })
    idioma = pandas.DataFrame(filas)
    buffer = io.BytesIO()
    with pandas.ExcelWriter(buffer, engine="openpyxl") as descarga:
        idioma.to_excel(descarga, index=False)
    buffer.seek(0)
    return buffer.getvalue()

def crearTablaCopiada(combinacionAsignaturas, sesionLabElegida):
    def agruparHuecos(filtroCuatrimestre):
        huecos = {}
        for fila in combinacionAsignaturas:
            cuatri = str(fila.get("Cuatrimestre", "1")).strip()
            if cuatri != filtroCuatrimestre:
                continue
            abrev = fila.get("Asignatura", "")
            grupo = fila.get("Grupo", "")
            labElegido = sesionLabElegida.get((abrev, grupo))

            horario_api_lab = st.session_state.get("horario_api_lab", None)

            if abrev == "API":
                for i in range(1, 5):
                    dia = normalizar(fila.get(f"T{i}-día", ""))
                    hora = normalizar(fila.get(f"T{i}-hora", ""))
                    if dia and hora:
                        huecos[(hora, dia)] = huecos.get((hora, dia), "") + f"{abrev}-T"
                if horario_api_lab:
                    cd, ch = horario_api_lab
                    huecos[(ch, cd)] = huecos.get((ch, cd), "") + f"{abrev}-L"
            else:
                for i in range(1, 5):
                    dia = normalizar(fila.get(f"T{i}-día", ""))
                    hora = normalizar(fila.get(f"T{i}-hora", ""))
                    if dia and hora:
                        huecos[(hora, dia)] = huecos.get((hora, dia), "") + f"{abrev}-T"
                for i in range(1, 5):
                    if labElegido is not None and labElegido != i:
                        continue
                    dia = normalizar(fila.get(f"L{i}-día", ""))
                    hora = normalizar(fila.get(f"L{i}-hora", ""))
                    if dia and hora:
                        huecos[(hora, dia)] = huecos.get((hora, dia), "") + f"{abrev}-L"
        return huecos

    lineas = []
    for cuatri, titulo in [("1", "1er Cuatrimestre"), ("2", "2o Cuatrimestre")]:
        huecos = agruparHuecos(cuatri)
        if not huecos:
            continue
        horas = sorted(set(hora for (hora, _) in huecos), key=convertirAMinutos)
        lineas.append(titulo)
        lineas.append("\t".join(["Hora"] + DIAS_ES))
        for hora in horas:
            celdaFila = [hora]
            for dia in DIAS_ES:
                celdaFila.append(huecos.get((hora, dia), ""))
            lineas.append("\t".join(celdaFila))
        lineas.append("")

    return "\n".join(lineas)

def cabecera(titulo, sub=""):
    subtitulo = (
        f"<div style='font-family:IBM Plex Mono,monospace;font-size:.74rem;"
        f"color:#888;margin-top:.25rem'>{sub}</div>"
        if sub else ""
    )
    st.markdown(
        f"<div style='background:#1c1c1c;color:#f0ede8;padding:1.6rem 2.2rem;"
        f"margin:-1rem -1rem 1.4rem -1rem;border-bottom:3px solid #e84545'>"
        f"<div style='font-family:IBM Plex Mono,monospace;font-size:1.5rem;"
        f"font-weight:700;letter-spacing:-.02em'>{titulo}</div>"
        f"{subtitulo}"
        f"</div>",
        unsafe_allow_html=True,
    )

###########################################################
# Las siguientes funciones serán empleadas para extraer el
# texto del PDF (text scraping) y generar un excel.
# Originalmente, tenía otro archivo .py aparte que se
# encargaba de hacer este proceso, pero la página tardaba
# en cargar, así que decidí juntarlos en el mismo .py.
###########################################################

def crearTabla(dataframe):
    resultado = defaultdict(lambda: {"Asignatura": "", "Curso": "", "Grupo": "", "T": [], "L": []})
    codigosIntens = {}
    filasAPI = []

    for _, fila in dataframe.iterrows():
        ident = fila["Identificador"]
        aulaT = normalizar(fila.get("AulaTeoria", ""))
        aulaL = normalizar(fila.get("AulaLab", "lab"))
        curso, _, grupo = ident.partition("\u00ba")
        curso = curso.strip()
        grupo = grupo.strip()
        hora = fila["Hora"].split("\n")[0]

        es_optativa = False
        for patron in PATRONESOPT:
            if re.search(patron, ident, re.IGNORECASE):
                es_optativa = True
                break
        
        intensificacionSeleccionada = None
        if not es_optativa:
            for abreviaturaIntensificacion in ["TI", "CO", "IS", "IC"]:
                if re.search(rf'[\[\( -]{abreviaturaIntensificacion}[\]\) -]', ident):
                    intensificacionSeleccionada = abreviaturaIntensificacion
                    ABREVIATURAS_INTESIFICACIONES.add(abreviaturaIntensificacion)
                    break

        entradasxfila = []
        for dia in DIAS:
            celda = fila.get(dia, "")
            if pandas.isna(celda) or str(celda).strip() == "":
                continue
            for b in str(celda).split("\n"):
                m = re.match(r"(.*?)-(t|p)", b.strip())
                if not m:
                    continue
                asig, tipo = m.groups()
                asig = asig.strip()
                entradasxfila.append((dia, asig, tipo))

        tieneAPI = any(asig == "API" for _, asig, _ in entradasxfila)

        for dia, asig, tipo in entradasxfila:
            if asig == "API":
                continue
            clave = (asig, curso, grupo)
            resultado[clave]["Asignatura"] = asig
            resultado[clave]["Curso"] = curso
            resultado[clave]["Grupo"] = grupo
            
            if intensificacionSeleccionada and not es_optativa and clave not in codigosIntens:
                if asig in TABLA_CODIGOS:
                    codigosIntens[clave] = intensificacionSeleccionada
            
            if tipo == "t":
                resultado[clave]["T"].append((dia, hora, aulaT))
            else:
                resultado[clave]["L"].append((dia, hora, aulaL))

        if tieneAPI:
            listado = {"hora": hora, "curso": curso, "T": [], "L": []}
            for dia, asig, tipo in entradasxfila:
                if asig != "API":
                    continue
                if tipo == "t" and dia in DIASTEORIAAPI:
                    listado["T"].append((dia, hora, aulaT))
                elif tipo == "p" and dia in DIASLABAPI:
                    listado["L"].append((dia, hora, aulaL))
            if listado["T"] or listado["L"]:
                filasAPI.append(listado)

    def crearFila(asig, curso, grupo, datos, intesificacion=None):
        bil, ef, ingles = obtenerBanderasIdioma(asig, TABLA_CODIGOS.get(asig))
        inglesSelec = bil or ef or ingles
        codigo_str = TABLA_CODIGOS.get(asig, "")
        coda = int(codigo_str) if codigo_str else ""
        info = get_infoAsignatura(coda) if coda else None
        
        intensificacion_final = ""
        if coda and coda in INTENSIFICACIONES:
            intensificacion_final = INTENSIFICACIONES[coda]
        elif intesificacion:
            intensificacion_final = intesificacion
        
        fila = {
            "Asignatura": asig, "Nombre ES": NOMBRES_ASIG.get(asig, asig),
            "Nombre EN": info["en"] if info else "",
            "Curso": curso, "Grupo": grupo,
            "Código": coda if coda else "",
            "Bilingue": bil, "EF": ef, "EN": inglesSelec,
            "Intensificacion": intensificacion_final,
        }
        for i in range(4):
            fila[f"T{i+1}-día"] = datos["T"][i][0] if len(datos["T"]) > i else ""
            fila[f"T{i+1}-hora"] = datos["T"][i][1] if len(datos["T"]) > i else ""
            fila[f"T{i+1}-aula"] = datos["T"][i][2] if len(datos["T"]) > i else ""
        for i in range(4):
            fila[f"L{i+1}-día"] = datos["L"][i][0] if len(datos["L"]) > i else ""
            fila[f"L{i+1}-hora"] = datos["L"][i][1] if len(datos["L"]) > i else ""
            fila[f"L{i+1}-aula"] = datos["L"][i][2] if len(datos["L"]) > i else ""
        return fila

    filas = []
    for (asig, curso, grupo), datos in resultado.items():
        filas.append(crearFila(asig, curso, grupo, datos,
                                intesificacion=codigosIntens.get((asig, curso, grupo))))

    horasDisp = set()
    for listado in filasAPI:
        hora = listado["hora"]
        if hora in horasDisp:
            continue
        horasDisp.add(hora)
        filas.append(crearFila("API", listado["curso"], hora,
                                {"T": listado["T"], "L": listado["L"]}))

    return pandas.DataFrame(filas)

def procesar_pdf(tamanio):
    try:
        grupo_tablas1 = get_tablas(tamanio, [6])
        grupo_tablas2 = get_tablas(tamanio, [7])
    except Exception:
        try:
            grupo_tablas1 = get_tablas(tamanio, [8])
            grupo_tablas2 = get_tablas(tamanio, [9])
        except Exception:
            grupo_tablas1 = get_tablas(tamanio, [7])
            grupo_tablas2 = get_tablas(tamanio, [8])
            
    t1 = crearTabla(grupo_tablas1)
    t1["Cuatrimestre"] = "1"
    t2 = crearTabla(grupo_tablas2)
    t2["Cuatrimestre"] = "2"
    combined = pandas.concat([t1, t2], ignore_index=True)
    return combined.drop_duplicates(subset=["Asignatura", "Grupo"]).reset_index(drop=True)

def cargar_excel(excel_bytes):
    excel = pandas.ExcelFile(io.BytesIO(excel_bytes))
    hojas = [hora for hora in excel.sheet_names if "correspon" not in hora.lower()]
    diasxhora = []
    for hora in hojas:
        dh = pandas.read_excel(excel, sheet_name=hora)
        if "Cuatrimestre" not in dh.columns:
            dh["Cuatrimestre"] = "2" if "segundo" in hora.lower() or hora.strip().endswith("2") else "1"
        diasxhora.append(dh)
    dataframe = pandas.concat(diasxhora, ignore_index=True)
    for col, valorPorDefecto in [("Bilingüe", False), ("EF", False), ("EN", False)]:
        if col not in dataframe.columns:
            dataframe[col] = valorPorDefecto
        dataframe[col] = dataframe[col].astype(bool)
    dataframe["Código"] = dataframe["Código"].apply(lambda v: convertirAEntero(v) or "")
    dataframe["Cuatrimestre"] = dataframe["Cuatrimestre"].astype(str).str.strip()
    
    def campoIngles(fila):
        if str(fila.get("Nombre EN", "")).strip() and str(fila.get("Nombre EN", "")).lower() != "nan":
            return fila["Nombre EN"]
        info = get_infoAsignatura(fila.get("Código"))
        return info["en"] if info else ""
    
    def campoEspaniol(fila):
        if str(fila.get("Nombre ES", "")).strip() and str(fila.get("Nombre ES", "")).lower() != "nan":
            return fila["Nombre ES"]
        return NOMBRES_ASIG.get(fila.get("Asignatura", ""), fila.get("Asignatura", ""))
    
    dataframe["Nombre EN"] = dataframe.apply(campoIngles, axis=1)
    dataframe["Nombre ES"] = dataframe.apply(campoEspaniol, axis=1)
    return dataframe.drop_duplicates(subset=["Asignatura", "Grupo"]).reset_index(drop=True)

def cargaDeIdiomas(excel_bytes):
    dataframe = pandas.read_excel(io.BytesIO(excel_bytes), header=None)
    filaCabecera = None
    for i, fila in dataframe.iterrows():
        if any("código grado" in str(v).lower() for v in fila.values):
            filaCabecera = i
            break
    if filaCabecera is None:
        raise ValueError("No se encontró la fila de encabezados en el Excel de idiomas.")
    dataframe.columns = dataframe.iloc[filaCabecera]
    dataframe = dataframe.iloc[filaCabecera + 1:].reset_index(drop=True)
    dataframe.columns = [str(c).strip() for c in dataframe.columns]
    codg_col = next((c for c in dataframe.columns if "código grado" in c.lower()), None)
    coda_col = next((c for c in dataframe.columns if "código asignatura" in c.lower()), None)
    espanicol = next((c for c in dataframe.columns if c.lower() == "español" or c.lower() == "español"), None)
    ef_col = next((c for c in dataframe.columns if "english friendly" in c.lower()), None)
    en_col = next((c for c in dataframe.columns if c.lower() == "inglés"), None)
    if not all([codg_col, coda_col, espanicol, ef_col, en_col]):
        raise ValueError(f"Columnas no encontradas. Encontradas: {list(dataframe.columns)}")
    dataframe = dataframe[dataframe[codg_col].apply(lambda v: str(v).strip()) == "406"].copy()
    sobreescritura = {}
    for _, fila in dataframe.iterrows():
        coda = convertirAEntero(fila[coda_col])
        if coda is None:
            continue
        def _yn(v):
            return str(v).strip().lower() == "si"
        sobreescritura[coda] = {
            "Español": _yn(fila[espanicol]),
            "EF": _yn(fila[ef_col]),
            "Inglés": _yn(fila[en_col]),
        }
    return sobreescritura

def sobreescribirIdiomas(dataframe, sobreescritura):
    if not sobreescritura:
        return dataframe
    dataframe = dataframe.copy()
    if "Bilingüe" in dataframe.columns and "Español" not in dataframe.columns:
        dataframe = dataframe.rename(columns={"Bilingüe": "Español"})
    for col in ["Español", "EF", "EN"]:
        if col not in dataframe.columns:
            dataframe[col] = False
    for indice, fila in dataframe.iterrows():
        coda = convertirAEntero(fila.get("Código"))
        if coda and coda in sobreescritura:
            dia = sobreescritura[coda]
            espaniol = bool(dia["Español"])
            ingles = bool(dia["Inglés"])
            ef = bool(dia["EF"])
            en = ingles
            dataframe.at[indice, "Español"] = espaniol
            dataframe.at[indice, "EF"] = ef
            dataframe.at[indice, "EN"] = en
    return dataframe

def generarHorarioHTML(celdas, titulo):
    DIAS = get_dias()
    if not celdas:
        return ""
    horas = sorted({hora for (_, hora) in celdas}, key=convertirAMinutos)
    html = '<div style="margin-bottom:1.2rem">'
    if titulo:
        html += f'<div style="font-family:IBM Plex Mono,monospace;font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.13em;color:#777;padding:.3rem 0 .4rem 0;border-bottom:1px solid #eee;margin-bottom:.3rem">{titulo}</div>'
    html += '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-family:IBM Plex Sans,sans-serif"><thead><tr style="background:#f5f4f0">'
    html += '<th style="padding:7px 10px;color:#999;font-size:.7rem;font-family:IBM Plex Mono,monospace;text-align:left;border-bottom:2px solid #eee;width:68px">Hora</th>'
    for dia in DIAS:
        html += f'<th style="padding:7px 10px;color:#1a1a1a;font-size:.8rem;font-weight:600;text-align:center;border-bottom:2px solid #eee">{dia}</th>'
    html += '<tr></thead><tbody>'
    for indice, hora in enumerate(horas):
        fondo = "#fff" if indice % 2 == 0 else "#fafaf8"
        html += f'<tr style="background:{fondo}"><td style="padding:5px 8px;color:#999;font-size:.7rem;font-family:IBM Plex Mono,monospace;border-right:1px solid #eee;white-space:nowrap;vertical-align:middle">{hora}'
        for d_disp in DIAS:
            d_es = DIAS_ES[DIAS_EN.index(d_disp)] if st.session_state.idioma == "en" else d_disp
            celda = celdas.get((d_es, hora), "")
            html += f'<td style="padding:4px 5px;border-right:1px solid #f0f0f0;vertical-align:top;min-width:120px">{celda}'
        html += '</tr>'
    html += '</tbody></table></div></div>'
    return html

def generarHorario(combinacionAsignaturas, colores, sesionLabElegida):
    cuatri1, cuatri2 = {}, {}
    teoria, laboratorio = obtenerTexto("teoria"), obtenerTexto("lab")
    gtapi = st.session_state.get("grupo_teoria_api", 1)
    glapi = st.session_state.get("grupo_lab_api", "")

    for fila in combinacionAsignaturas:
        abrev = fila.get("Asignatura", "")
        grupo = fila.get("Grupo", "")
        cuatri = normalizar(fila.get("Cuatrimestre", "1")) or "1"
        fondo, texto = colores.get(abrev, ("#eee", "#333"))
        cuatri = cuatri2 if cuatri == "2" else cuatri1
        labElegido = sesionLabElegida.get((abrev, grupo))
        disp = mostrarNombre(abrev, fila.get("Código"))

        if abrev == "API":
            for i in range(1, 5):
                dia = normalizar(fila.get(f"T{i}-día", ""))
                hora = normalizar(fila.get(f"T{i}-hora", ""))
                aula = normalizar(fila.get(f"T{i}-aula", ""))
                if dia and hora:
                    set_celda(cuatri, dia, hora, fondo, texto, disp, aula, "✍︎", teoria)

            horario_api_lab = st.session_state.get("horario_api_lab", None)
            if horario_api_lab:
                dia, hora = horario_api_lab
                api = st.session_state.get("df_asignaturas")
                aula = "lab"
                if api is not None:
                    for _, filaAPI in api[api["Asignatura"] == "API"].iterrows():
                        for i in range(1, 5):
                            if (normalizar(filaAPI.get(f"L{i}-día", "")) == dia and
                                normalizar(filaAPI.get(f"L{i}-hora", "")) == hora):
                                aula = normalizar(filaAPI.get(f"L{i}-aula", "")) or "lab"
                                break
                set_celda(cuatri, dia, hora, fondo, texto, disp, aula, "⌨", laboratorio)
            else:
                for i in range(1, 5):
                    dia = normalizar(fila.get(f"L{i}-día", ""))
                    hora = normalizar(fila.get(f"L{i}-hora", ""))
                    aula = normalizar(fila.get(f"L{i}-aula", "")) or "lab"
                    if dia and hora:
                        set_celda(cuatri, dia, hora, fondo, texto, disp, aula, "⌨", laboratorio)
        else:
            for i in range(1, 5):
                dia = normalizar(fila.get(f"T{i}-día", ""))
                hora = normalizar(fila.get(f"T{i}-hora", ""))
                aula = normalizar(fila.get(f"T{i}-aula", ""))
                if dia and hora:
                    set_celda(cuatri, dia, hora, fondo, texto, disp, aula, "✍︎", teoria)

            for i in range(1, 5):
                if labElegido is not None and labElegido != i:
                    continue
                dia = normalizar(fila.get(f"L{i}-día", ""))
                hora = normalizar(fila.get(f"L{i}-hora", ""))
                aula = normalizar(fila.get(f"L{i}-aula", "")) or "lab"
                if dia and hora:
                    set_celda(cuatri, dia, hora, fondo, texto, disp, aula, "⌨", laboratorio)

    tieneDos = bool(cuatri2)
    html = generarHorarioHTML(cuatri1, obtenerTexto("cuatri1_etiqueta") if tieneDos else "")
    html += generarHorarioHTML(cuatri2, obtenerTexto("cuatri2_etiqueta")) if tieneDos else ""
    resumen = '<div style="display:flex;flex-wrap:wrap;gap:6px;margin-top:.8rem;padding-top:.6rem;border-top:1px solid #eee">'
    for fila in combinacionAsignaturas:
        abrev = fila.get("Asignatura", "")
        fondo, texto = colores.get(abrev, ("#eee", "#333"))
        resumen += (f'<span style="background:{fondo};color:{texto};border:1px solid {texto};border-radius:3px;' +
                    f'padding:2px 9px;font-size:.76rem;font-weight:600;font-family:IBM Plex Mono,monospace">{abrev}</span>')
    resumen += '</div>'
    return html + resumen

###########################################################
# En este segmento he incorporado el funcionamiento de la
# barra lateral, la cual permite cambiar entre el modo
# Erasmus y el modo de estudiante común, además de 
# seleccionar la página de horario o de fuente de datos, la
# cual posibilita cargar el nuevo POD y los idiomas de las
# asignaturas, así como modificar estas en caso de haber
# algún error en el POD (en este de 2026-2027 pone TWS en
# lugar de TSW para la intensificación de TI, por lo que el
# programa lo identifica como otra asignatura).
###########################################################

with st.sidebar:
    st.markdown(f"""<div style='padding:1.2rem 0 .6rem 0'>
        <div style='font-family:IBM Plex Mono,monospace;font-size:.96rem;font-weight:700;color:#f0ede8;letter-spacing:-.01em'>{obtenerTexto("titulo")}</div>
        <div style='font-family:IBM Plex Mono,monospace;font-size:.58rem;color:#666;text-transform:uppercase;letter-spacing:.1em;margin-top:.2rem'>{obtenerTexto("subtitulo")}</div>
        </div><div style='border-top:1px solid #2e2e2e;margin-bottom:.6rem'></div>""",
                unsafe_allow_html=True)
    columnaIzquierda1, columnaIzquierda2 = st.columns(2)
    with columnaIzquierda1:
        if st.button("🇪🇸 ES", key="b_es", type="primary" if st.session_state.idioma == "es" else "secondary"):
            st.session_state.idioma = "es"
            st.rerun()
    with columnaIzquierda2:
        if st.button("🇬🇧 EN", key="b_en", type="primary" if st.session_state.idioma == "en" else "secondary"):
            st.session_state.idioma = "en"
            st.rerun()
    st.markdown("<div style='border-top:1px solid #2e2e2e;margin:.55rem 0 .65rem 0'></div>", unsafe_allow_html=True)
    
    for clave, etiqueta in [("horario", obtenerTexto("generar_horario")), ("fuenteDatos", obtenerTexto("fuente_de_datos"))]:
        if st.button(etiqueta, key=f"nav_{clave}", use_container_width=True,
                     type="primary" if st.session_state.seccion == clave else "secondary"):
            st.session_state.seccion = clave
            st.rerun()
    
    if st.session_state.df_asignaturas is not None:
        num = len(st.session_state.df_asignaturas)
        st.markdown(f"<div style='border-top:1px solid #2e2e2e;margin-top:.75rem;padding-top:.65rem'><div style='font-family:IBM Plex Mono,monospace;font-size:.66rem;color:#4ade80'>✔ {num} {obtenerTexto('asigs')}</div></div>",
                    unsafe_allow_html=True)
    intensificacionActiva = primeraIntensificacion()
    if intensificacionActiva:
        st.markdown(f"<div style='font-family:IBM Plex Mono,monospace;font-size:.64rem;color:#c084fc;margin-top:.3rem'>⬡ {NOMBRES_INTENSIFICACIONES.get(intensificacionActiva, intensificacionActiva)}</div>",
                    unsafe_allow_html=True)

    st.markdown("<div style='border-top:1px solid #2e2e2e;margin-top:.7rem;padding-top:.7rem'></div>", unsafe_allow_html=True)
    etiquetaErasmus = "✈ Erasmus" if not st.session_state.erasmus else "✈ Erasmus ✔"
    tipoErasmus = "primary" if st.session_state.erasmus else "secondary"
    if st.button(etiquetaErasmus, key="btn_erasmus", use_container_width=True, type=tipoErasmus):
        estaDesactivado = not st.session_state.erasmus
        st.session_state.erasmus = estaDesactivado
        if not estaDesactivado:
            selec = st.session_state.asignaturas_seleccionadas
            if selec and st.session_state.df_asignaturas is not None:
                aux = st.session_state.df_asignaturas
                intensificacionSeleccionada = None
                asignaturasEliminadas = []
                for claveAbreviaturas in list(selec.keys()):
                    claveFilas = aux[aux["Asignatura"] == claveAbreviaturas]
                    if claveFilas.empty:
                        continue
                    intensificacionActiva = get_intensificacionPorFila(claveFilas.iloc[0])
                    if intensificacionActiva is None:
                        continue
                    if intensificacionSeleccionada is None:
                        intensificacionSeleccionada = intensificacionActiva
                    elif intensificacionActiva != intensificacionSeleccionada:
                        asignaturasEliminadas.append(claveAbreviaturas)
                for claveAbreviaturas in asignaturasEliminadas:
                    del selec[claveAbreviaturas]
                    st.session_state.idiomas_seleccionados.pop(claveAbreviaturas, None)
        st.rerun()

###########################################################
# Aquí configuro el correcto funcionamiento de la página
# que sirve para cargar la fuente de datos, es decir, la
# encargada de leer el POD y generar un archivo excel en la
# propia página web, así como cargar el excel de idiomas y
# configurar el excel con los idiomas correctos que incluye
# este archivo excel.
###########################################################

if st.session_state.seccion == "fuenteDatos":
    cabecera(obtenerTexto("titulo_admin"), obtenerTexto("administrar_datos"))

    if not st.session_state.modo_admin:
        st.markdown("<br>", unsafe_allow_html=True)
        aplicarCSSACajaTexto(obtenerTexto("acceso"))
        st.markdown("<br>", unsafe_allow_html=True)
        columnaIzquierda, columnaCentral, columnaDerecha = st.columns([1, 2, 1])
        with columnaCentral:
            pswd = st.text_input(obtenerTexto("pswd"), type="password", key="pwd_field", label_visibility="visible")
            if st.button(obtenerTexto("desbloquear"), type="primary", use_container_width=True):
                if pswd == PSWD:
                    st.session_state.modo_admin = True
                    st.rerun()
                else:
                    textoComodin("Error", obtenerTexto("pswd_incorrecta"))
        st.stop()

    cargarPOD, cargarIdiomas, editarTabla, exportarTabla = st.tabs([
        obtenerTexto("cargar_pod"), obtenerTexto("tabla_idiomas"), obtenerTexto("editar_tabla"), obtenerTexto("exportar_tabla")
    ])

    with cargarPOD:
        cargar = st.file_uploader(obtenerTexto("etiqueta_carga"), type=["pdf", "xlsx"], key="uploader")
        if cargar:
            datos = cargar.read()
            extension = cargar.name.lower().split(".")[-1]
            with st.spinner(obtenerTexto("cargando")):
                try:
                    cargaNueva = (procesar_pdf(datos) if extension == "pdf" else cargar_excel(datos))
                    cargaNueva = sobreescribirIdiomas(cargaNueva, st.session_state.lang_overrides)
                    st.session_state.df_asignaturas = cargaNueva.copy()
                    st.session_state.asignaturas_seleccionadas = {}
                    st.session_state.idiomas_seleccionados = {}
                    st.session_state.sesiones_practicas = {}
                    guardar_cambios(cargaNueva, st.session_state.lang_overrides)
                    textoCorrecto(f"✔ {obtenerTexto('pod_cargado') if extension == 'pdf' else obtenerTexto('excel_cargado')} ({len(cargaNueva)} {obtenerTexto('filas')})")
                except Exception as e:
                    textoComodin(obtenerTexto("error_carga"), str(e))

    with cargarIdiomas:
        aplicarCSSACajaTexto(obtenerTexto("pista_idioma"))
        if st.session_state.lang_overrides:
            idiomas = crearExcelIdiomas(st.session_state.lang_overrides)
            st.download_button(
                f"⬇ {obtenerTexto('descargar_idiomas')}",
                idiomas,
                "idiomas_actuales.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_lang_current"
            )
        st.markdown("<br>", unsafe_allow_html=True)
        cargaIdiomas = st.file_uploader(obtenerTexto("cargar_idiomas"), type=["xlsx"], key="lang_uploader")
        if cargaIdiomas:
            try:
                sobreescritura = cargaDeIdiomas(cargaIdiomas.read())
                st.session_state.lang_overrides = sobreescritura
                if st.session_state.df_asignaturas is not None:
                    st.session_state.df_asignaturas = sobreescribirIdiomas(
                        st.session_state.df_asignaturas, sobreescritura)
                guardar_cambios(st.session_state.df_asignaturas, sobreescritura)
                num = len(sobreescritura)
                textoCorrecto(f"✔ {obtenerTexto('idiomas_cargados')} ({num} asignaturas actualizadas)")
                filas = []
                for coda, dia in sobreescritura.items():
                    info = get_infoAsignatura(coda)
                    filas.append({"Código": coda,
                                 "Nombre ES": info["es"] if info else str(coda),
                                 "🇪🇸 Español": dia["Español"],
                                 "EF": dia["EF"],
                                 "🇬🇧 Inglés": dia["Inglés"]})
                st.dataframe(pandas.DataFrame(filas), use_container_width=True, hide_index=True)
            except Exception as e:
                textoComodin(obtenerTexto("error_idiomas"), str(e))

    with editarTabla:
        tablaSinProcesar = st.session_state.df_asignaturas
        if tablaSinProcesar is None:
            aplicarCSSACajaTexto(obtenerTexto("sin_datos"))
        else:
            tablaProcesada = tablaSinProcesar.drop_duplicates(subset=["Asignatura", "Grupo"]).reset_index(drop=True).copy()
            columnasPrincipales = ["Asignatura", "Nombre ES", "Nombre EN", "Curso", "Grupo", "Código", "Español", "EF", "EN", "Cuatrimestre"]
            columnasSecundarias = [c for c in tablaProcesada.columns if c not in columnasPrincipales and c not in ["Bilingüe"]]
            if "Bilingüe" in tablaProcesada.columns and "Español" not in tablaProcesada.columns:
                tablaProcesada = tablaProcesada.rename(columns={"Bilingüe": "Español"})
            tablaProcesada = tablaProcesada[[c for c in columnasPrincipales if c in tablaProcesada.columns] + columnasSecundarias]
            st.markdown(f"<div style='font-size:.82rem;color:#666;margin-bottom:.4rem'>{len(tablaProcesada)} {obtenerTexto('filas')} · {obtenerTexto('ordenar')}</div>", unsafe_allow_html=True)
            edicion = st.data_editor(tablaProcesada, use_container_width=True, num_rows="dynamic", key="editor",
                column_config={
                    "Código": st.column_config.NumberColumn("Código", format="%d"),
                    "Curso": st.column_config.NumberColumn("Curso", format="%d"),
                    "Español": st.column_config.CheckboxColumn("🇪🇸 Español"),
                    "EF": st.column_config.CheckboxColumn("EF"),
                    "EN": st.column_config.CheckboxColumn("🇬🇧 EN"),
                })
            if st.button("⬇ " + obtenerTexto("guardar_cambios"), type="primary"):
                st.session_state.df_asignaturas = edicion.copy()
                guardar_cambios(edicion, st.session_state.lang_overrides)
                textoCorrecto(f"✔ {obtenerTexto('cambios_guardados')}")

    with exportarTabla:
        tablaSinProcesar = st.session_state.df_asignaturas
        if tablaSinProcesar is None:
            aplicarCSSACajaTexto(obtenerTexto("sin_datos"))
        else:
            st.markdown(f"<div style='font-size:.82rem;color:#666;margin-bottom:.6rem'>{len(tablaSinProcesar)} {obtenerTexto('filas')}</div>", unsafe_allow_html=True)
            st.dataframe(tablaSinProcesar, use_container_width=True)
            st.markdown("<br>", unsafe_allow_html=True)
            descargar, copiar = st.columns(2)
            with descargar:
                buffer = io.BytesIO()
                with pandas.ExcelWriter(buffer, engine="openpyxl") as descarga:
                    tablaSinProcesar.to_excel(descarga, index=False)
                st.download_button(f"⬇ {obtenerTexto('exportar_datos')}", buffer.getvalue(), "pod.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with copiar:
                botonCopiar(obtenerTexto("copiar_al_portapapeles"), copiarTexto(tablaSinProcesar), "ds_main")

###########################################################
# Al igual que he hecho en el anterior fragmento de código,
# ahora he configurado el correcto funcionamiento de la
# página que permite crear un horario por medio de las
# asignaturas encontradas en el POD. Además, esta página es
# la página que aparece al ingresar en la página, lo que
# resulta mucho más agradable y reduce el tiempo de
# elaboración del horario.
###########################################################

elif st.session_state.seccion == "horario":
    cabecera(obtenerTexto("titulo_horario"))
    dataframe = st.session_state.df_asignaturas
    if dataframe is None:
        aplicarCSSACajaTexto(obtenerTexto("no_datos"))
        st.stop()

    with st.expander("🔍︎ " + ("Filtros" if st.session_state.idioma == "es" else "Filters"), expanded=True):
        filtroCurso, filtroCuatri, filtroIdioma = st.columns(3)
        with filtroCurso:
            st.markdown(f"**{obtenerTexto('anio')}**")
            cursos = sorted(dataframe["Curso"].dropna().astype(str).unique().tolist())
            cursoSeleccionado = {c: st.checkbox(f"{obtenerTexto('anio')} {c}", value=False, key=f"ck_c_{c}") for c in cursos}
            curso_f = [c for c, v in cursoSeleccionado.items() if v] or cursos
        with filtroCuatri:
            st.markdown(f"**{obtenerTexto('cuatri')}**")
            cuatri1 = st.checkbox(obtenerTexto("cuatri1"), value=False, key="ck_q1")
            cuatri2 = st.checkbox(obtenerTexto("cuatri2"), value=False, key="ck_q2")
            cuatri_f = (["1"] if cuatri1 else []) + (["2"] if cuatri2 else []) or ["1", "2"]
        with filtroIdioma:
            st.markdown(f"**{obtenerTexto('idioma')}**")
            bilingue = st.checkbox(obtenerTexto("bil"), value=False, key="ck_bil")
            englishFriendly = st.checkbox(obtenerTexto("ef"), value=False, key="ck_ef")

    filtros = dataframe[dataframe["Curso"].astype(str).isin(curso_f) & dataframe["Cuatrimestre"].astype(str).isin(cuatri_f)].copy()

    estaEspaniol = "Español" in filtros.columns
    estaEF = "EF" in filtros.columns
    estaIngles = "EN" in filtros.columns

    if bilingue and not englishFriendly:
        if estaEspaniol and estaIngles:
            filtros = filtros[(filtros["Español"] == True) & (filtros["EN"] == True)]
        elif estaEspaniol:
            filtros = filtros[filtros["Español"] == True]
    elif englishFriendly and not bilingue:
        if estaEF:
            filtros = filtros[filtros["EF"] == True]
    elif bilingue and englishFriendly:
        filtrar = pandas.Series([False] * len(filtros), index=filtros.index)
        if estaEspaniol and estaIngles:
            filtrar = filtrar | ((filtros["Español"] == True) & (filtros["EN"] == True))
        if estaEF:
            filtrar = filtrar | (filtros["EF"] == True)
        filtros = filtros[filtrar]

    def ordenarClave(fila):
        info = get_infoAsignatura(fila.get("Código"))
        nombre = (info["en"] if st.session_state.idioma == "en" else info["es"]) if info else str(fila.get("Asignatura", ""))
        return nombre.upper() + "|||" + str(fila.get("Grupo", ""))
    
    filtros = filtros.copy()
    if not filtros.empty:
        filtros["_sk"] = filtros.apply(ordenarClave, axis=1)
        filtros = filtros.sort_values("_sk").drop(columns=["_sk"]).reset_index(drop=True)
    
    asignaturas = filtros.drop_duplicates(subset=["Asignatura"]).reset_index(drop=True) if not filtros.empty else filtros

    if filtros.empty:
        nadaEncontrado = "No hay asignaturas que coincidan con los filtros seleccionados." if st.session_state.idioma == "es" else "No subjects match the selected filters."
        aplicarCSSACajaTexto(nadaEncontrado)
        st.stop()

    elegidas = list(st.session_state.asignaturas_seleccionadas.items())
    solapes = comprobarSolapes(elegidas, dataframe, st.session_state.sesiones_practicas)
    abreviaturasConflictivas = {c[0][0] for c in solapes} | {c[1][0] for c in solapes}
    intensificacionElegida = primeraIntensificacion()
    erasmus = st.session_state.get("erasmus", False)

    aplicarCSSASeccion(f"{len(asignaturas)} {obtenerTexto('asig_disp')} — {obtenerTexto('pista')}")

    consulta = st.text_input("", placeholder=obtenerTexto("texto_busqueda"), key="subject_search", label_visibility="collapsed")
    if consulta:
        textoDeBusqueda = consulta.strip().lower()
        def _matches(row_a):
            if textoDeBusqueda in str(row_a.get("Asignatura", "")).lower():
                return True
            if textoDeBusqueda in str(row_a.get("Código", "")).lower():
                return True
            if textoDeBusqueda in str(row_a.get("Nombre ES", "")).lower():
                return True
            if textoDeBusqueda in str(row_a.get("Nombre EN", "")).lower():
                return True
            return False
        asignaturas = asignaturas[asignaturas.apply(_matches, axis=1)].reset_index(drop=True)
        if len(asignaturas) == 1 and len(textoDeBusqueda) >= 2:
            unicaAsig = asignaturas.iloc[0]
            unicaAbrev = unicaAsig["Asignatura"]
            if unicaAbrev not in st.session_state.asignaturas_seleccionadas:
                grupos_auto = sorted(dataframe[dataframe["Asignatura"] == unicaAbrev]["Grupo"].dropna().astype(str).unique().tolist())
                st.session_state.asignaturas_seleccionadas[unicaAbrev] = grupos_auto[0] if grupos_auto else "auto"
                st.rerun()

    for _, filaAsignatura in asignaturas.iterrows():
        abrev = filaAsignatura["Asignatura"]
        codigo = filaAsignatura.get("Código", None)
        coda = convertirAEntero(codigo)
        info = get_infoAsignatura(coda)
        bil, ef, ingles = obtenerBanderasIdioma(abrev, coda)
        banderaInglesa = bil or ef or ingles
        muestraNombre = mostrarNombre(abrev, coda)
        nombreTraducido = (info["en"] if st.session_state.idioma == "es" else info["es"]) if info else ""
        estaSeleccionada = abrev in st.session_state.asignaturas_seleccionadas
        haySolape = abrev in abreviaturasConflictivas and estaSeleccionada
        asignaturaDeIntensificacion = get_intensificacionPorFila(filaAsignatura)

        bloqueoIntensificacion = False
        if asignaturaDeIntensificacion and intensificacionElegida and asignaturaDeIntensificacion != intensificacionElegida and not estaSeleccionada and not erasmus:
            bloqueoIntensificacion = True

        colorBorde = "#d63030" if haySolape else ("#1c1c1c" if estaSeleccionada else ("#aaa" if bloqueoIntensificacion else "#ddd"))
        estiloBorde = f"border-left:3px solid {colorBorde};" if (estaSeleccionada or haySolape) else ""
        fondoCelda = "#fff2f2" if haySolape else ("#f8f7f4" if estaSeleccionada else ("#f5f5f5" if bloqueoIntensificacion else "#fff"))
        opacidad = "opacity:.5;" if bloqueoIntensificacion else ""

        etiquetasAsignatura = ""
        if bil:
            etiquetasAsignatura += "<span class='b-es'>🇪🇸 ES</span> "
        if banderaInglesa:
            etiquetasAsignatura += "<span class='b-en'>EN</span> "
        if ef:
            etiquetasAsignatura += "<span class='b-ef'>EF</span> "
        if asignaturaDeIntensificacion:
            etiquetasAsignatura += f"<span class='b-track'>{asignaturaDeIntensificacion}</span> "
        if not bil and not ef and not banderaInglesa:
            etiquetasAsignatura += "<span class='b-es'>🇪🇸 ES</span> "

        codaS = str(coda) if coda else "—"
        
        with st.container():
            columnas = st.columns([5, 2, 3])
            with columnas[0]:
                st.markdown(f"""<div style='border:1px solid {colorBorde};{estiloBorde}border-radius:5px;
                    background:{fondoCelda};padding:.6rem .9rem;margin-bottom:3px;{opacidad}'>
                    <div style='display:flex;align-items:flex-start;gap:.5rem'>
                        <span style='font-family:IBM Plex Mono,monospace;font-size:.68rem;color:#aaa;min-width:44px;padding-top:2px'>{codaS}</span>
                        <div style='flex:1'>
                            <div style='font-size:.88rem;font-weight:700;color:{"#d63030" if haySolape else "#1a1a1a"}'>
                                {"ⓘ " if haySolape else ""}{"✔ " if estaSeleccionada else ""}{muestraNombre}</div>
                            <div style='font-size:.73rem;color:#888;font-style:italic;margin-top:1px'>
                                {nombreTraducido if nombreTraducido and nombreTraducido.upper() != muestraNombre.upper() else ""}</div>
                        </div>
                        <div style='white-space:nowrap'>{etiquetasAsignatura}</div>
                    </div></div>""", unsafe_allow_html=True)

            with columnas[1]:
                if estaSeleccionada:
                    opciones = []
                    if bil:
                        opciones += [obtenerTexto("idioma_es"), obtenerTexto("idioma_en")]
                    if ef:
                        opciones += [obtenerTexto("idioma_ef")]
                    if not opciones:
                        opciones = [obtenerTexto("idioma_es")]
                    idiomaActual = st.session_state.idiomas_seleccionados.get(abrev, opciones[0])
                    if idiomaActual not in opciones:
                        idiomaActual = opciones[0]
                    if len(opciones) > 1:
                        for opcion in opciones:
                            if st.checkbox(opcion, value=(idiomaActual == opcion), key=f"lck_{abrev}_{opcion}"):
                                st.session_state.idiomas_seleccionados[abrev] = opcion
                    else:
                        st.markdown(f"<div style='font-size:.77rem;color:#777;padding-top:.5rem'>{opciones[0]}</div>", unsafe_allow_html=True)
                        st.session_state.idiomas_seleccionados[abrev] = opciones[0]

            with columnas[2]:
                if estaSeleccionada:
                    grupos = sorted(dataframe[dataframe["Asignatura"] == abrev]["Grupo"].dropna().astype(str).unique().tolist())
                    grupoActual = st.session_state.asignaturas_seleccionadas.get(abrev, grupos[0] if grupos else "")
                    if grupoActual not in grupos and grupos:
                        grupoActual = grupos[0]
                    seleccionarGrupo = st.selectbox(obtenerTexto("grp"), grupos, index=grupos.index(grupoActual) if grupoActual in grupos else 0,
                                    key=f"grp_{abrev}", label_visibility="collapsed")
                    st.session_state.asignaturas_seleccionadas[abrev] = seleccionarGrupo
                    if st.button(obtenerTexto("eliminar"), key=f"rm_{abrev}", type="secondary"):
                        if abrev in st.session_state.asignaturas_seleccionadas:
                            del st.session_state.asignaturas_seleccionadas[abrev]
                        st.session_state.idiomas_seleccionados.pop(abrev, None)
                        st.rerun()
                else:
                    if bloqueoIntensificacion:
                        st.markdown(f"<div style='font-size:.74rem;color:#888;padding-top:.5rem'>{NOMBRES_INTENSIFICACIONES.get(intensificacionElegida, intensificacionElegida)}</div>", unsafe_allow_html=True)
                    else:
                        if st.button(obtenerTexto("add"), key=f"add_{abrev}", type="primary"):
                            grupos = sorted(dataframe[dataframe["Asignatura"] == abrev]["Grupo"].dropna().astype(str).unique().tolist())
                            st.session_state.asignaturas_seleccionadas[abrev] = grupos[0] if grupos else "auto"
                            st.rerun()

    st.markdown("---")
    if not st.session_state.asignaturas_seleccionadas:
        st.stop()
    
    seleccion = st.session_state.asignaturas_seleccionadas
    seleccionarAsignaturas = list(seleccion.keys())
    aplicarCSSASeccion(f"{len(seleccionarAsignaturas)} {obtenerTexto('seleccion')}")
    
    for abrev, grupo in seleccion.items():
        filaAsignaturaActual = dataframe[dataframe["Asignatura"] == abrev]
        coda = convertirAEntero(filaAsignaturaActual.iloc[0]["Código"]) if not filaAsignaturaActual.empty else None
        idioma = st.session_state.idiomas_seleccionados.get(abrev, "")
        haySolape = abrev in abreviaturasConflictivas
        etiquetaGrupo = grupo if grupo != "auto" else obtenerTexto("grp_aleatorio")
        disp = mostrarNombre(abrev, coda)
        st.markdown(f"""<div style='display:flex;align-items:center;gap:.5rem;padding:.28rem 0;border-bottom:1px solid #eee'>
            <span style='font-family:IBM Plex Mono,monospace;font-size:.66rem;color:#aaa;width:44px'>{str(coda) if coda else "—"}</span>
            <span style='font-size:.85rem;font-weight:700;color:{"#d63030" if haySolape else "#1a1a1a"};flex:1'>{"ⓘ " if haySolape else ""}{disp}</span>
            <span style='font-size:.76rem;color:#666'>{obtenerTexto("grp")} {etiquetaGrupo}</span>
            {"<span style='font-size:.73rem;color:#888;margin-left:.4rem'>"+idioma+"</span>" if idioma else ""}
            </div>""", unsafe_allow_html=True)

    if st.session_state.asignaturas_seleccionadas:
        st.markdown("<br>", unsafe_allow_html=True)
        columnaIzquierda, columnaCentral, columnaDerecha = st.columns([3, 1, 3])
        with columnaCentral:
            if st.button("✘ " + obtenerTexto("borrar_todo"), key="btn_clear_schedule", type="secondary", use_container_width=True):
                st.session_state.asignaturas_seleccionadas = {}
                st.session_state.idiomas_seleccionados = {}
                st.session_state.sesiones_practicas = {}
                st.session_state.indice_api_lab = 0
                st.session_state.horario_api_lab = None
                st.rerun()

    elegidas = list(seleccion.items())
    solapes = comprobarSolapes(elegidas, dataframe, st.session_state.sesiones_practicas)
    if solapes:
        st.markdown("<br>", unsafe_allow_html=True)
        for s in solapes:
            a1, g1 = s[0]
            a2, g2 = s[1]
            cadenaSolapes = ", ".join(f"{traducirDia(dia)} {hora}" for dia, hora in s[2])
            textoComodin(obtenerTexto("titulo_solapes"), f"<strong>{mostrarNombre(a1)}</strong> (G{g1 if g1 != 'auto' else '?'}) {obtenerTexto('solape')}: <strong>{cadenaSolapes}</strong> ↔ <strong>{mostrarNombre(a2)}</strong> (G{g2 if g2 != 'auto' else '?'})")
        columna1, columna2 = st.columns([3, 1])
        with columna1:
            st.markdown(f"<div style='font-size:.83rem;color:#555;padding-top:.4rem'>{obtenerTexto('consejo_solape')}</div>", unsafe_allow_html=True)
        with columna2:
            st.button(obtenerTexto("omitir_solapes"), type="secondary")
    else:
        textoCorrecto(f"✔ {obtenerTexto('no_solapes')}")

    aplicarCSSASeccion(obtenerTexto("sesion"))
    st.markdown(f"<div style='font-size:.81rem;color:#555;margin-bottom:.7rem'>{obtenerTexto('pista_lab')}</div>", unsafe_allow_html=True)
    colores = {abrev: COLORES[i % len(COLORES)] for i, abrev in enumerate(seleccionarAsignaturas)}
    tieneGruposPracticas = False

    for abrev in seleccionarAsignaturas:
        grupo = seleccion[abrev]
        filas = dataframe[(dataframe["Asignatura"] == abrev) & (dataframe["Grupo"].astype(str) == str(grupo))]
        if filas.empty:
            continue
        fila = filas.iloc[0]
        fondo, texto = colores.get(abrev, ("#eee", "#333"))

        if abrev == "API":
            filasAPI = dataframe[dataframe["Asignatura"] == "API"]
            gruposLabAPI = []
            huecosLabAPI = set()
            for _, filaAPI in filasAPI.iterrows():
                for i in range(1, 5):
                    dia = normalizar(filaAPI.get(f"L{i}-día", ""))
                    hora = normalizar(filaAPI.get(f"L{i}-hora", ""))
                    aula = normalizar(filaAPI.get(f"L{i}-aula", "")) or "lab"
                    if dia and hora and (dia, hora) not in huecosLabAPI:
                        huecosLabAPI.add((dia, hora))
                        gruposLabAPI.append((f"{traducirDia(dia)} {hora} 🖈{aula}", (dia, hora)))
            if len(gruposLabAPI) == 0:
                continue
            tieneGruposPracticas = True
            etiquetasPracticas = [lbl for lbl, _ in gruposLabAPI]
            practicaSeleccionada = st.session_state.get("indice_api_lab", 0)
            practicaSeleccionada = min(practicaSeleccionada, len(etiquetasPracticas) - 1)
            columnaIzquierda1, columnaIzquierda2 = st.columns([2, 5])
            with columnaIzquierda1:
                st.markdown(f"<span style='font-weight:700;font-size:.84rem;color:{texto};background:{fondo};padding:2px 8px;border-radius:3px'>API</span>", unsafe_allow_html=True)
            with columnaIzquierda2:
                eleccionLabAPI = st.radio("", etiquetasPracticas, index=practicaSeleccionada, key="api_lab_radio", label_visibility="collapsed")
                indiceAPILabSeleccionado = etiquetasPracticas.index(eleccionLabAPI)
                st.session_state["indice_api_lab"] = indiceAPILabSeleccionado
                st.session_state["horario_api_lab"] = gruposLabAPI[indiceAPILabSeleccionado][1]
        else:
            diaPractica1, horaPractica1 = normalizar(fila.get("L1-día")), normalizar(fila.get("L1-hora"))
            diaPractica2, horaPractica2 = normalizar(fila.get("L2-día")), normalizar(fila.get("L2-hora"))
            if not (diaPractica1 and horaPractica1 and diaPractica2 and horaPractica2):
                continue
            tieneGruposPracticas = True
            aulaPractica1 = normalizar(fila.get("L1-aula", "lab"))
            aulaPractica2 = normalizar(fila.get("L2-aula", "lab"))
            infoCompletaPractica1 = f"{obtenerTexto('sesion1')}: {traducirDia(diaPractica1)} {horaPractica1}" + (f" — 🖈 {aulaPractica1}" if aulaPractica1 else "")
            infoCompletaPractica2 = f"{obtenerTexto('sesion2')}: {traducirDia(diaPractica2)} {horaPractica2}" + (f" — 🖈 {aulaPractica2}" if aulaPractica2 else "")
            practicaActual = st.session_state.sesiones_practicas.get((abrev, grupo), 1)
            columnaIzquierda1, columnaIzquierda2 = st.columns([2, 5])
            with columnaIzquierda1:
                st.markdown(f"<span style='font-weight:700;font-size:.84rem;color:{texto};background:{fondo};padding:2px 8px;border-radius:3px'>{abrev}</span>", unsafe_allow_html=True)
            with columnaIzquierda2:
                selec = st.radio("", options=[1, 2], index=0 if practicaActual == 1 else 1,
                               format_func=lambda x: infoCompletaPractica1 if x == 1 else infoCompletaPractica2,
                               key=f"lab_radio_{abrev}_{grupo}", horizontal=True,
                               label_visibility="collapsed")
                st.session_state.sesiones_practicas[(abrev, grupo)] = selec

    if not tieneGruposPracticas:
        st.markdown(f"<div style='font-size:.81rem;color:#aaa'>{obtenerTexto('no_lab')}</div>", unsafe_allow_html=True)

    aplicarCSSASeccion(obtenerTexto("etiqueta_horario"))
    combinacionAsignaturas = []
    for abrev, grupo in seleccion.items():
        filas = dataframe[(dataframe["Asignatura"] == abrev) & (dataframe["Grupo"].astype(str) == str(grupo))]
        if not filas.empty:
            r = filas.iloc[0].to_dict()
            r["Grupo"] = grupo
            combinacionAsignaturas.append(r)
    horarioIntegroEnPagina = generarHorario(combinacionAsignaturas, colores, st.session_state.sesiones_practicas)
    num = len(combinacionAsignaturas)
    sufijo = "subject(s)" if st.session_state.idioma == "en" else "asignatura(s)"
    st.markdown(
        f"<div class='tt-wrap'>"
        f"<div class='tt-hdr'>{obtenerTexto('etiqueta_horario')} — {num} {sufijo}</div>"
        f"<div style='padding:1rem'>{horarioIntegroEnPagina}</div>"
        f"</div>",
        unsafe_allow_html=True)

    aplicarCSSASeccion(obtenerTexto("leyenda"))
    leyenda = "<div style='display:flex;flex-wrap:wrap;gap:7px'>"
    for abrev in seleccionarAsignaturas:
        fondo, texto = colores.get(abrev, ("#eee", "#333"))
        grupo = seleccion.get(abrev, "")
        idioma = st.session_state.idiomas_seleccionados.get(abrev, "")
        filaAsignaturaActual = dataframe[dataframe["Asignatura"] == abrev]
        coda = convertirAEntero(filaAsignaturaActual.iloc[0]["Código"]) if not filaAsignaturaActual.empty else None
        info = get_infoAsignatura(coda)
        nombreEspaniol = info["es"] if info else abrev
        nombreIngles = info["en"] if info else abrev
        idiomaPrimario = nombreIngles if st.session_state.idioma == "en" else nombreEspaniol
        idiomaSecundario = nombreEspaniol if st.session_state.idioma == "en" else nombreIngles
        leyenda += (f"<div style='background:{fondo};border:1px solid {texto};border-radius:4px;padding:5px 10px'>"
                f"<div style='font-family:IBM Plex Mono,monospace;font-size:.68rem;font-weight:700;color:{texto}'>{abrev}</div>"
                f"<div style='font-size:.74rem;color:#1a1a1a;margin-top:1px'>{idiomaPrimario}</div>"
                f"<div style='font-size:.66rem;color:#888;font-style:italic'>{idiomaSecundario if idiomaSecundario != idiomaPrimario else ''}</div>"
                f"<div style='font-size:.66rem;color:#666'>G{grupo}{(' · '+idioma) if idioma else ''}</div></div>")
    leyenda += "</div>"
    st.markdown(leyenda, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    filas = []
    for fila in combinacionAsignaturas:
        abrev = fila.get("Asignatura", "")
        grupo = fila.get("Grupo", "")
        coda = convertirAEntero(fila.get("Código", ""))
        info = get_infoAsignatura(coda)
        fila = {"Asignatura": abrev, "Nombre ES": info["es"] if info else abrev, "Nombre EN": info["en"] if info else abrev,
                "Grupo": grupo, "Idioma/Language": st.session_state.idiomas_seleccionados.get(abrev, ""),
                "Lab session": st.session_state.sesiones_practicas.get((abrev, str(grupo)), "—")}
        for col in ["Código", "Curso", "Cuatrimestre", "T1-día", "T1-hora", "T1-aula", "T2-día", "T2-hora", "T2-aula", "L1-día", "L1-hora", "L1-aula", "L2-día", "L2-hora", "L2-aula"]:
            fila[col] = fila.get(col, "")
        filas.append(fila)
    exportarTabla = pandas.DataFrame(filas)
    descargarTablaEnExcel, descargarTextoEnExcel, copiarInfoTabla = st.columns(3)
    with descargarTablaEnExcel:
        tamanioTabla = crearTablaExcel(combinacionAsignaturas, st.session_state.sesiones_practicas, st.session_state.lang_overrides)
        st.download_button(f"⬇ {obtenerTexto('descargar_horario')}", tamanioTabla, "mi_horario_tabla.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with descargarTextoEnExcel:
        buffer = io.BytesIO()
        with pandas.ExcelWriter(buffer, engine="openpyxl") as descarga:
            exportarTabla.to_excel(descarga, index=False)
        st.download_button(f"⬇ {obtenerTexto('descargar_horario_excel')}", buffer.getvalue(), "mi_horario_datos.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with copiarInfoTabla:
        infoAlPortapapeles = crearTablaCopiada(combinacionAsignaturas, st.session_state.sesiones_practicas)
        botonCopiar(obtenerTexto("copiar_horario"), infoAlPortapapeles, "sched")